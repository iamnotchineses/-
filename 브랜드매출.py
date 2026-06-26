import html
import io
import re
from pathlib import Path

import numpy as np
import pandas as pd
import plotly.express as px
import plotly.graph_objects as go
import streamlit as st

st.set_page_config(page_title="메카 매출 대시보드", page_icon="📊", layout="wide")

APP_DIR = Path(__file__).parent


def find_default_data() -> Path | None:
    """업로드가 없을 때 사용할 기본 데이터 파일을 앱 폴더에서 자동 선택한다.
    1순위: '완성_sales_revenue*.xlsx' 중 가장 최신(수정시각 → 파일명 역순)
    2순위: 기존 샘플 'sample_mecca_raw.xlsx'
    둘 다 없으면 None.
    """
    candidates = sorted(
        APP_DIR.glob("완성_sales_revenue*.xlsx"),
        key=lambda p: (p.stat().st_mtime, p.name),
        reverse=True,
    )
    if candidates:
        return candidates[0]
    legacy = APP_DIR / "sample_mecca_raw.xlsx"
    return legacy if legacy.exists() else None


SAMPLE_FILE = find_default_data()

# -----------------------------
# Style
# -----------------------------
st.markdown(
    """
    <style>
    .main .block-container {padding-top: 1.5rem; padding-bottom: 2rem;}
    div[data-testid="stMetric"] {
        background: #ffffff;
        border: 1px solid #e8eef5;
        padding: 14px 14px;
        border-radius: 16px;
        box-shadow: 0 4px 16px rgba(15, 23, 42, 0.04);
    }
    div[data-testid="stMetric"] label {color:#64748b; font-size:0.85rem;}
    div[data-testid="stMetricValue"] {
        font-size: 1.3rem;
        font-weight: 700;
        white-space: nowrap;
        overflow: hidden;
        text-overflow: ellipsis;
    }
    div[data-testid="stMetricValue"] > div {
        overflow: hidden;
        text-overflow: ellipsis;
        white-space: nowrap;
    }
    .section-title {
        font-size: 1.15rem;
        font-weight: 800;
        margin: 1.2rem 0 .4rem 0;
        color: #0f172a;
    }
    .hint {
        color: #64748b;
        font-size: .9rem;
    }
    .printtbl { border-collapse: collapse; width: 100%; font-size: 12px; margin: 6px 0; }
    .printtbl th, .printtbl td { border: 1px solid #d0d7de; padding: 4px 8px; text-align: right; white-space: nowrap; }
    .printtbl th { background: #f1f5f9; text-align: center; font-weight: 700; }
    .printtbl td.l { text-align: left; }
    .printtbl tbody tr:nth-child(even) { background: #fafbfc; }
    @media print {
        @page { size: A4 landscape; margin: 8mm; }
        [data-testid="stSidebar"], [data-testid="stToolbar"], [data-testid="stHeader"],
        [data-testid="stDecoration"], header, footer, [data-testid="stStatusWidget"] { display: none !important; }
        .stApp, .main, .block-container { background: #fff !important; padding-top: 0 !important; }
        section.main div.block-container, .main .block-container, .block-container { max-width: 100% !important; }
        * { -webkit-print-color-adjust: exact !important; print-color-adjust: exact !important; }
        /* 차트 폭만 용지에 맞춤 (레이아웃은 화면 그대로 둠) */
        [data-testid="stPlotlyChart"], [data-testid="stPlotlyChart"] > div,
        .js-plotly-plot, .js-plotly-plot .main-svg, .svg-container, .plot-container {
            width: 100% !important; max-width: 100% !important;
        }
        /* 차트/표 한 줄만 잘리지 않게 */
        [data-testid="stPlotlyChart"], .js-plotly-plot { page-break-inside: avoid; }
        .printtbl tr { page-break-inside: avoid; }
    }
    </style>
    """,
    unsafe_allow_html=True,
)

# -----------------------------
# Helpers
# -----------------------------
def clean_col_name(col: str) -> str:
    return re.sub(r"\s+", " ", str(col).replace("\n", " ")).strip()


def _xlsx_rows_fast(data: bytes) -> list:
    """openpyxl read_only + data_only 로 첫 시트를 값만 빠르게 읽음.
    임베드 이미지/수식이 많은 무거운 .xlsx 도 빠르고 메모리 적게 읽는다."""
    from openpyxl import load_workbook
    wb = load_workbook(io.BytesIO(data), read_only=True, data_only=True)
    try:
        ws = wb.worksheets[0]
        rows = [list(r) for r in ws.iter_rows(values_only=True)]
    finally:
        wb.close()
    while rows and all(c is None for c in rows[-1]):  # 빈 꼬리행 제거
        rows.pop()
    return rows


def _html_rows_fast(text: str) -> list:
    """HTML 위장 .xls 의 <table> 를 lxml 스트리밍 파싱(<tr> 단위로 읽고 해제 → 100MB+도 메모리 적게)."""
    from lxml import etree
    rows = []
    ctx = etree.iterparse(io.BytesIO(text.encode("utf-8")), events=("end",),
                          tag="tr", html=True, recover=True, encoding="utf-8")
    for _, tr in ctx:
        cells = []
        for cell in tr:
            if isinstance(cell.tag, str) and cell.tag in ("td", "th"):
                txt = "".join(cell.itertext()).strip()
                cells.append(txt if txt != "" else None)
        if cells:
            rows.append(cells)
        tr.clear()
        while tr.getprevious() is not None:  # 처리한 행 메모리 해제
            del tr.getparent()[0]
    return rows


def _dedupe_cols(cols):
    seen, out = {}, []
    for c in cols:
        c = c if (c is not None and str(c).strip() != "") else "col"
        if c in seen:
            seen[c] += 1; out.append(f"{c}.{seen[c]}")
        else:
            seen[c] = 0; out.append(c)
    return out


def read_excel_smart(file_obj) -> pd.DataFrame:
    """Read Excel where the first row may be blank and the actual header starts later.
    openpyxl read_only(값만)로 빠르게 읽고, 실패 시 pandas 로 폴백."""
    data = file_obj.read() if hasattr(file_obj, "read") else file_obj
    try:
        rows = _xlsx_rows_fast(data)
    except Exception:
        raw = pd.read_excel(io.BytesIO(data), sheet_name=0, header=None)
        rows = [list(r) for r in raw.where(pd.notna(raw), None).values.tolist()]
    if not rows:
        return pd.DataFrame()
    keywords = {"주문번호", "쇼핑몰", "브랜드", "수량", "최종판매가", "출고날짜"}
    header_row = 0
    for i in range(min(10, len(rows))):
        values = set(str(x).strip() for x in rows[i] if x is not None)
        if len(values & keywords) >= 3:
            header_row = i
            break
    cols = _dedupe_cols([str(c).strip() if c is not None else "" for c in rows[header_row]])
    df = pd.DataFrame(rows[header_row + 1:], columns=cols, dtype=object)
    df.columns = [clean_col_name(c) for c in df.columns]
    # 엑셀 AA열(27번째 = 0-based 26)을 '위치 그대로' 확보(헤더 비면 Unnamed 로 잡혀 삭제되므로)
    aa_series = df.iloc[:, 26].copy() if df.shape[1] > 26 else None
    df = df.loc[:, ~pd.Series(df.columns).astype(str).str.startswith("Unnamed").values]
    df = df.dropna(how="all")
    if aa_series is not None and "대분류" not in df.columns:
        df["대분류"] = aa_series.reindex(df.index)
    return df


def find_col(df: pd.DataFrame, candidates: list[str], fallback_contains: str | None = None) -> str | None:
    cols = list(df.columns)
    for c in candidates:
        if c in cols:
            return c
    if fallback_contains:
        for c in cols:
            if fallback_contains in c:
                return c
    return None


def to_number(s: pd.Series) -> pd.Series:
    if s is None:
        return pd.Series(dtype=float)
    return pd.to_numeric(
        s.astype(str).str.replace(",", "", regex=False).str.replace("%", "", regex=False),
        errors="coerce",
    ).fillna(0)


def _parse_date_flexible(series: pd.Series) -> pd.Series:
    """날짜 안전 파싱. 엑셀 serial(정수)·YYYYMMDD·다양한 포맷 문자열·datetime 객체 혼재 모두 처리.
    숫자를 무조건 epoch(나노초)로 보던 문제(45292/20240715/0 → 1970-01 → FW69) 방지."""
    if series is None or len(series) == 0:
        return pd.Series(pd.NaT, index=getattr(series, "index", None), dtype="datetime64[ns]")
    if pd.api.types.is_datetime64_any_dtype(series):
        return pd.to_datetime(series, errors="coerce")
    out = pd.Series(pd.NaT, index=series.index, dtype="datetime64[ns]")
    # 1) 이미 날짜/시간 객체 (openpyxl datetime 등)
    is_dt = series.apply(lambda v: (not isinstance(v, str)) and hasattr(v, "year"))
    if is_dt.any():
        out.loc[is_dt] = pd.to_datetime(series[is_dt], errors="coerce")
    # 2) 숫자: YYYYMMDD(8자리) vs 엑셀 serial 구분
    num = pd.to_numeric(series.where(~is_dt), errors="coerce")
    ymd = num.between(19000101, 21001231)
    if ymd.any():
        out.loc[ymd] = pd.to_datetime(
            num[ymd].round().astype("int64").astype(str), format="%Y%m%d", errors="coerce")
    ser = num.between(20000, 60000) & ~ymd  # 약 1954~2064년 엑셀 날짜 serial
    if ser.any():
        out.loc[ser] = pd.to_datetime(num[ser], unit="D", origin="1899-12-30", errors="coerce")
    # 3) 나머지 문자열 (2024-07-15, 2024.7.15, 2024/07/15, 시간 포함 등 포맷 혼재 가능)
    _txt = series.astype(str).str.strip()
    rest = (~is_dt) & num.isna() & _txt.ne("") & ~_txt.str.lower().isin(["nan", "nat", "none"])
    if rest.any():
        try:
            out.loc[rest] = pd.to_datetime(_txt[rest], errors="coerce", format="mixed")
        except (ValueError, TypeError):
            out.loc[rest] = pd.to_datetime(_txt[rest], errors="coerce")
    return out


def money(v) -> str:
    try:
        return f"{float(v):,.0f}원"
    except Exception:
        return "0원"


def num(v) -> str:
    try:
        return f"{float(v):,.0f}"
    except Exception:
        return "0"


def eok(v) -> str:
    """금액 표기: 1억 이상이면 'X.X억', 그 미만이면 전체 숫자(콤마)."""
    try:
        v = float(v)
    except Exception:
        return "0"
    if pd.isna(v):
        return "-"
    if abs(v) >= 1e8:
        return f"{v / 1e8:.1f}억"
    return f"{v:,.0f}"


def to_line(model) -> str:
    """모델명에서 끝의 사이즈 '(...)' 를 떼어 라인명으로 변환.
    예: 'COHBU M26388 ALI BLANC/BLEU CIEL (XL)' -> 'COHBU M26388 ALI BLANC/BLEU CIEL'
        'K100979-001 (44)' -> 'K100979-001'
    """
    return re.sub(r"\s*\([^()]*\)\s*$", "", str(model)).strip()


line_map: dict = {}  # 모델명(정규화) → 라인명. 로드 시 (이미지 3번째 시트 + 재고)로 채움.


def _norm_model(s) -> str:
    return re.sub(r"\s+", " ", str(s)).strip()


def _line_of(model) -> str:
    """모델명 → 라인명. line_map(이미지 3번째 시트/재고) 우선, 없으면 끝의 사이즈 '(...)' 제거."""
    nm = _norm_model(model)
    if nm in line_map:
        return line_map[nm]
    nm2 = _norm_model(to_line(model))
    if nm2 in line_map:
        return line_map[nm2]
    return to_line(model)


def pct(v) -> str:
    try:
        if pd.isna(v) or np.isinf(v):
            return "-"
        return f"{float(v):,.1f}%"
    except Exception:
        return "-"


def growth_pct(v) -> str:
    """Format growth-rate columns like ▲ 18.0% / ▼ 15.0% with one decimal."""
    try:
        if pd.isna(v) or np.isinf(v):
            return "-"
        value = float(v)
        if value > 0:
            return f"▲ {abs(value):,.1f}%"
        if value < 0:
            return f"▼ {abs(value):,.1f}%"
        return "0.0%"
    except Exception:
        return "-"


def add_rate(df: pd.DataFrame, current_col: str, prev_col: str, out_col: str = "YoY 신장률") -> pd.DataFrame:
    prev = df[prev_col].replace(0, np.nan)
    df[out_col] = ((df[current_col] - df[prev_col]) / prev.abs()) * 100
    return df


# ===== 인쇄/PDF 출력 지원 (메인 대시보드와 동일) =====
PRINT_MODE = False  # 사이드바에서 켜면 표를 정적 HTML(인쇄/PDF용)로 렌더

_MONEY_KW = ["판매가", "매출", "수익원", "원가", "증감", "객단가", "수수료액", "배송비",
             "신장액", "총매출", "총원가", "금액", "목표", "실제", "재고", "입고원가"]


def _is_money_col(c) -> bool:
    """금액(콤마) 컬럼인지. %(률/율/비중)는 제외."""
    c = str(c)
    if any(k in c for k in ("률", "율", "비중", "Rate", "달성", "소진율", "회수율")):
        return False
    return (any(k in c for k in _MONEY_KW)
            or bool(re.fullmatch(r"\d{4}", c))
            or bool(re.fullmatch(r"\d{4}년", c)))


def _fmt_cell(col, val) -> str:
    """인쇄용 HTML 표 셀 표시값(숫자 콤마/%/부호)."""
    if pd.isna(val):
        return "-"
    cs = str(col)
    if not isinstance(val, (int, float, np.integer, np.floating)):
        return str(val)  # 이미 문자열(▲▼, 시즌, 라인명 등)
    try:
        if cs in ("Rank", "순위", "#"):
            return f"{int(val)}"
        if "신장률" in cs or "신장율" in cs or "대비" in cs:
            return f"{val:+.1f}%"
        if any(k in cs for k in ["률", "율", "비중", "Rate", "달성"]):
            return f"{val:.1f}%"
        if cs in ("수량", "주문수", "라인수", "재고수량"):
            return f"{val:,.0f}"
        if _is_money_col(cs):
            return f"{val:,.0f}"
    except Exception:
        pass
    return str(val)


def _df_to_html(df: pd.DataFrame) -> str:
    """데이터프레임 → 인쇄 친화 정적 HTML 표."""
    name_cols = {"쇼핑몰", "브랜드", "대분류", "모델명", "요일", "공식/병행", "라인명", "시즌", "연도", "분기"}
    head = "".join(f"<th>{html.escape(str(c))}</th>" for c in df.columns)
    body = []
    for _, r in df.iterrows():
        tds = []
        for c in df.columns:
            cls = " class='l'" if str(c) in name_cols else ""
            tds.append(f"<td{cls}>{html.escape(_fmt_cell(c, r[c]))}</td>")
        body.append("<tr>" + "".join(tds) + "</tr>")
    return (f"<table class='printtbl'><thead><tr>{head}</tr></thead>"
            f"<tbody>{''.join(body)}</tbody></table>")


def render_table(df: pd.DataFrame, **kwargs) -> None:
    """PRINT_MODE면 정적 HTML 표, 아니면 일반 표(정렬 가능)."""
    if PRINT_MODE:
        st.markdown(_df_to_html(df), unsafe_allow_html=True)
    else:
        st.dataframe(df, **kwargs)


def sort_desc(df: pd.DataFrame, by: str) -> pd.DataFrame:
    if by in df.columns:
        return df.sort_values(by=by, ascending=False, na_position="last")
    return df


# 표/차트의 '총매출' 라벨 접두어 — 본문에서 모드(주간/월간)에 따라 재설정됨
INTERVAL_LABEL = "주간"


def format_table(df: pd.DataFrame):
    """표시용 (DataFrame, column_config) 반환.
    금액/수량은 '숫자'로 유지해 콤마 표시(localized)와 숫자 정렬이 둘 다 되게 한다.
    비율(수익률/비중)은 숫자+'%' 포맷, 증감률(신장률)만 ▲▼ 문자열, Rank는 문자열.
    """
    out = df.copy()
    # 수량을 최종판매가(총매출) 바로 왼쪽으로 이동
    if "수량" in out.columns and "최종판매가" in out.columns:
        cols = list(out.columns)
        cols.remove("수량")
        cols.insert(cols.index("최종판매가"), "수량")
        out = out[cols]
    rename_map = {}
    if "최종판매가" in out.columns:
        rename_map["최종판매가"] = "총매출"
    if "수익원(실배송비)" in out.columns:
        rename_map["수익원(실배송비)"] = "수익원"
    if rename_map:
        out = out.rename(columns=rename_map)
    money_keywords = ["판매가", "매출", "수익원", "원가", "증감", "객단가", "수수료액", "배송비", "신장액"]
    pct_keywords = ["률", "율", "비중", "Rate"]
    colcfg = {}
    for c in out.columns:
        cs = str(c)
        if c == "Rank":
            out[c] = pd.to_numeric(out[c], errors="coerce").apply(lambda x: "-" if pd.isna(x) else f"{x:,.0f}")
        elif "신장률" in cs or "신장율" in cs:
            out[c] = pd.to_numeric(out[c], errors="coerce").apply(growth_pct)
        elif any(k in cs for k in pct_keywords):
            out[c] = pd.to_numeric(out[c], errors="coerce").replace([np.inf, -np.inf], np.nan)
            colcfg[c] = st.column_config.NumberColumn(cs, format="%.1f%%")
        elif any(k in cs for k in money_keywords) or re.fullmatch(r"\d{4}년", cs) or re.fullmatch(r"\d{1,2}월\s?\d{1,2}주차", cs):
            out[c] = pd.to_numeric(out[c], errors="coerce").round(0).astype("Int64")
            colcfg[c] = st.column_config.NumberColumn(cs, format="localized")
        elif cs in ("수량", "주문수"):
            out[c] = pd.to_numeric(out[c], errors="coerce").round(0).astype("Int64")
            colcfg[c] = st.column_config.NumberColumn(cs, format="localized")
        else:
            out[c] = out[c].replace({None: "-", np.nan: "-"})
    return out, colcfg


def aggregate(df: pd.DataFrame, group_cols: list[str], metric_cols: dict) -> pd.DataFrame:
    agg_spec = {}
    for out, col in metric_cols.items():
        if col and col in df.columns:
            agg_spec[out] = (col, "sum")
    result = df.groupby(group_cols, dropna=False).agg(**agg_spec).reset_index()
    if "최종판매가" in result.columns and "수량" in result.columns:
        result["객단가"] = np.where(result["수량"] != 0, result["최종판매가"] / result["수량"], 0)
    if "수익원(실배송비)" in result.columns and "최종판매가" in result.columns:
        result["수익률"] = np.where(result["최종판매가"] != 0, result["수익원(실배송비)"] / result["최종판매가"] * 100, 0)
    total = result["최종판매가"].sum() if "최종판매가" in result.columns else 0
    if total != 0 and "최종판매가" in result.columns:
        result["매출비중"] = result["최종판매가"] / total * 100
    return sort_desc(result, "최종판매가")



def rank_table(df: pd.DataFrame, name_col: str) -> pd.DataFrame:
    """Add rank and replace the displayed name with rank order prefix."""
    out = df.copy().reset_index(drop=True)
    out.insert(0, "Rank", np.arange(1, len(out) + 1))
    if name_col in out.columns:
        clean_name = out[name_col].astype(str).str.replace(r"^\s*\d+\s*[\.\)\-_/]*\s*", "", regex=True)
        out[name_col] = out["Rank"].astype(str) + ". " + clean_name
    return out


def top_sales_table(df: pd.DataFrame, group_cols: list[str], topn: int = 30, sort_by: str = "최종판매가") -> pd.DataFrame:
    table = aggregate(df, group_cols, metric_cols)
    if sort_by in table.columns:
        table = table.sort_values(sort_by, ascending=False, na_position="last")
    table = table.head(topn).reset_index(drop=True)
    table.insert(0, "Rank", np.arange(1, len(table) + 1))
    return table

def yoy_by_group(df: pd.DataFrame, group_col: str, base_year: int, metric_col: str) -> pd.DataFrame:
    prev_year = base_year - 1
    temp = df[df["연도"].isin([prev_year, base_year])]
    pivot = temp.pivot_table(index=group_col, columns="연도", values=metric_col, aggfunc="sum", fill_value=0).reset_index()
    if prev_year not in pivot.columns:
        pivot[prev_year] = 0
    if base_year not in pivot.columns:
        pivot[base_year] = 0
    pivot = pivot.rename(columns={prev_year: f"{prev_year}년", base_year: f"{base_year}년"})
    pivot["YoY 신장액"] = pivot[f"{base_year}년"] - pivot[f"{prev_year}년"]
    pivot = add_rate(pivot, f"{base_year}년", f"{prev_year}년")
    return sort_desc(pivot, f"{base_year}년")


def trend_by_group(df: pd.DataFrame, group_col: str, metric_col: str, topn=None) -> pd.DataFrame:
    """그룹별 연도 추이: 데이터에 존재하는 모든 연도를 열로 펼친다. 최신 연도 기준 내림차순."""
    years = sorted(int(y) for y in df["연도"].dropna().unique())
    pivot = df.pivot_table(
        index=group_col, columns="연도", values=metric_col, aggfunc="sum", fill_value=0
    ).reset_index()
    for y in years:
        if y not in pivot.columns:
            pivot[y] = 0
    pivot = pivot.rename(columns={y: f"{y}년" for y in years})
    year_cols = [f"{y}년" for y in years]
    pivot = pivot[[group_col] + year_cols]
    if year_cols:
        pivot = pivot.sort_values(year_cols[-1], ascending=False, na_position="last")
    pivot = pivot.reset_index(drop=True)
    if topn:
        pivot = pivot.head(topn)
    return pivot


def wow_by_group(df: pd.DataFrame, group_col: str, metric_col: str, week_order: list, topn=None) -> pd.DataFrame:
    """그룹별 주차 추이: week_order(시간순 주차 라벨)대로 열을 펼친다. 최신 주차 기준 내림차순."""
    pivot = df.pivot_table(
        index=group_col, columns="주차", values=metric_col, aggfunc="sum", fill_value=0
    ).reset_index()
    week_cols = [w for w in week_order if w in pivot.columns]
    pivot = pivot[[group_col] + week_cols]
    if week_cols:
        pivot = pivot.sort_values(week_cols[-1], ascending=False, na_position="last")
    pivot = pivot.reset_index(drop=True)
    if topn:
        pivot = pivot.head(topn)
    return pivot


@st.cache_data(show_spinner=False)
def load_data_from_bytes(data: bytes | None) -> pd.DataFrame:
    if data is None:
        default_path = find_default_data()
        if default_path is None:
            raise FileNotFoundError(
                "기본 데이터 파일을 찾지 못했습니다. 왼쪽에서 엑셀을 업로드하거나, "
                "app.py 와 같은 폴더에 '완성_sales_revenue*.xlsx' 파일을 두세요."
            )
        with open(default_path, "rb") as f:
            data = f.read()
    return _finalize_df(read_excel_smart(io.BytesIO(data)))


@st.cache_data(show_spinner="데이터 처리 중… (첫 로드는 행 수에 따라 수십 초 걸릴 수 있어요)")
def load_upload(raw: bytes) -> pd.DataFrame:
    """원본 bytes → (원본이면 가공) → 최종 DataFrame.
    가공 경로는 xlsx 직렬화/재파싱을 생략하고 메모리에서 바로 DataFrame 을 만든다(대용량 속도 핵심)."""
    rows = _parse_raw_rows(raw)
    if _is_already_processed(rows):
        return _finalize_df(read_excel_smart(io.BytesIO(raw)))  # 완성형은 그대로
    col_names, data_rows = _process_raw_rows(rows)
    df = pd.DataFrame(data_rows, columns=col_names, dtype=object)
    return _finalize_df(df)


def _finalize_df(df: pd.DataFrame) -> pd.DataFrame:
    df.columns = [clean_col_name(c) for c in df.columns]

    # Detect columns
    date_col = find_col(df, ["출고날짜", "출고일", "판매일자", "주문일자"], "날짜")
    qty_col = find_col(df, ["수량", "판매수량"], "수량")
    gross_col = find_col(df, ["매출가", "총매출액", "매출액"], "매출")
    net_col = find_col(df, ["최종판매가", "순매출액", "실매출액"], "최종")
    profit_col = find_col(df, ["수익원(실배송비)", "수익원 실배송비", "수익원", "공헌이익"], "수익원")
    cost_col = find_col(df, ["원가총액", "출고원가"], "원가")
    mall_col = find_col(df, ["쇼핑몰", "몰", "채널"], "쇼핑몰")
    brand_col = find_col(df, ["브랜드", "브랜드명"], "브랜드")
    # 대분류: 위에서 AA열을 '대분류'로 보존했다. G열 '대카테고리'(브랜드패션 등)는 절대 쓰지 않는다.
    if "대분류" in df.columns:
        category_col = "대분류"
    else:
        category_col = find_col(df, ["카테고리", "분류"], "분류")
    model_col = find_col(df, ["모델명", "상품명", "품목명", "상품코드"], "모델")
    order_col = find_col(df, ["주문번호", "주문ID", "주문코드"], "주문")
    note_col = find_col(df, ["비고", "상태", "구분"], "비고")

    # 공식/병행 구분 컬럼: 헤더명 우선, 없으면 값이 공식/병행 으로만 이루어진 컬럼을 자동 탐색
    official_col = find_col(df, ["공식/병행", "공식병행", "공식여부"])
    if official_col is None:
        for _c in df.columns:
            _vals = set(df[_c].dropna().astype(str).str.strip().unique())
            if _vals and _vals <= {"공식", "병행"}:
                official_col = _c
                break

    # Standardize important columns
    if date_col is None:
        raise ValueError("날짜 컬럼을 찾지 못했습니다. '출고날짜' 또는 날짜가 포함된 컬럼이 필요합니다.")
    df["날짜"] = pd.to_datetime(df[date_col], errors="coerce")
    df = df.dropna(subset=["날짜"]).copy()
    # 반품(수량<0): 비고의 '반품(YYYY-MM-DD)' 실제 반품일을 출고날짜/분석날짜로 사용 (출고일 혼동 방지)
    if qty_col and note_col and qty_col in df.columns and note_col in df.columns:
        _ret = pd.to_numeric(df[qty_col], errors="coerce").fillna(0) < 0
        if _ret.any():
            _rd = pd.to_datetime(
                df.loc[_ret, note_col].astype(str).str.extract(r"(\d{4}[-./]\d{1,2}[-./]\d{1,2})", expand=False),
                errors="coerce")
            df.loc[_ret, "날짜"] = _rd.fillna(df.loc[_ret, "날짜"])
            if date_col and date_col in df.columns:
                df.loc[_ret, date_col] = df.loc[_ret, "날짜"]
    df["연도"] = df["날짜"].dt.year.astype(int)
    df["월"] = df["날짜"].dt.month.astype(int)
    df["연월"] = df["날짜"].dt.to_period("M").astype(str)
    # (주간/월간 기간 파생은 아래 수량 표준화 후에 수행 — 반품=수량 음수 제외 위해)
    df["요일순"] = df["날짜"].dt.weekday.astype(int)
    df["요일라벨"] = df["요일순"].map({0: "월", 1: "화", 2: "수", 3: "목", 4: "금", 5: "토", 6: "일"})

    for col in [qty_col, gross_col, net_col, profit_col, cost_col]:
        if col and col in df.columns:
            df[col] = to_number(df[col])

    if qty_col and qty_col != "수량":
        df["수량"] = df[qty_col]
    elif "수량" not in df.columns:
        df["수량"] = 0

    # (주간/월간/연간 기간 파생[주시작·주차]은 본문에서 '합쳐진 전체 df' 기준으로 수행한다.
    #  멀티파일 업로드 시 파일마다 따로 잡혀 라벨/모드가 틀어지는 문제 방지)

    if gross_col and gross_col != "매출가":
        df["매출가"] = df[gross_col]
    elif "매출가" not in df.columns:
        df["매출가"] = 0

    if net_col and net_col != "최종판매가":
        df["최종판매가"] = df[net_col]
    elif "최종판매가" not in df.columns:
        df["최종판매가"] = df["매출가"]

    # 매출가 컬럼이 비어있는(거의 0인) export 에서는 최종판매가를 매출가로 사용한다.
    if df["매출가"].abs().sum() == 0 or (df["매출가"] != 0).mean() < 0.05:
        df["매출가"] = df["최종판매가"]

    if profit_col and profit_col != "수익원(실배송비)":
        df["수익원(실배송비)"] = df[profit_col]
    elif "수익원(실배송비)" not in df.columns:
        df["수익원(실배송비)"] = 0

    if cost_col and cost_col != "원가총액":
        df["원가총액"] = df[cost_col]
    elif "원가총액" not in df.columns:
        df["원가총액"] = 0

    for std, col in {
        "쇼핑몰": mall_col,
        "브랜드": brand_col,
        "대분류": category_col,
        "공식/병행": official_col,
        "모델명": model_col,
        "주문번호": order_col,
        "비고": note_col,
    }.items():
        if col and col in df.columns:
            df[std] = df[col].fillna("미분류").astype(str)
        elif std not in df.columns:
            df[std] = "미분류"
        else:
            df[std] = df[std].fillna("미분류").astype(str)

    # Normalize text values
    for c in ["쇼핑몰", "브랜드", "대분류", "공식/병행", "모델명", "비고"]:
        df[c] = df[c].replace({"nan": "미분류", "None": "미분류", "": "미분류"})

    # 대분류 정규화: 세부 카테고리(드레스/모자/상의 등)를 _CATEGORY_MAP 으로 8개 대분류에 통합.
    #   8개(시계/주얼리/가방/지갑/의류/신발/소품/용품) 밖이면 '미분류'. 원본은 진단용 보존.
    try:
        df["대분류_원본"] = df["대분류"].astype(str)
        df["대분류"] = df["대분류"].map(_classify_to8)
    except Exception:
        pass

    return df


@st.cache_data(show_spinner=False)
def _img_keys(name) -> set:
    """이미지 매칭 키 후보: 원본 + 괄호 사이즈 제거형(카드의 to_line 매칭용)."""
    s = "" if name is None else str(name).strip()
    out = {s}
    out.add(re.sub(r"\s*\([^()]*\)\s*$", "", s).strip())
    return {k for k in out if k}


@st.cache_data(show_spinner=False)
def load_image_map_from_bytes(data: bytes | None) -> dict:
    """엑셀의 두 번째(옆) 시트에서 A열=라인명, B열=이미지URL 매핑을 읽는다.
    시트가 없거나 형식이 안 맞으면 빈 dict 를 반환(이미지 없이 동작)."""
    if data is None:
        return {}
    try:
        xls = pd.ExcelFile(io.BytesIO(data))
    except Exception:
        return {}
    other_sheets = list(xls.sheet_names[1:])  # 첫 시트(메인 데이터) 제외
    mapping: dict[str, str] = {}
    for sh in other_sheets:
        try:
            sub = pd.read_excel(xls, sheet_name=sh, header=None, usecols=[0, 1])
        except Exception:
            continue
        for _, row in sub.iterrows():
            name = row.iloc[0]
            url = str(row.iloc[1]).strip() if pd.notna(row.iloc[1]) else ""
            if pd.isna(name) or not url.lower().startswith(("http://", "https://")):
                continue  # 헤더행/빈값/비 URL 스킵
            for k in _img_keys(name):
                mapping[k] = url
    return mapping


def _norm_mall(s) -> str:
    """쇼핑몰명 매칭용 정규화: 공백/괄호/'주식회사' 제거, 소문자."""
    s = "" if s is None else str(s)
    s = re.sub(r"\s+", "", s).lower()
    for t in ("주식회사", "(", ")", "（", "）"):
        s = s.replace(t, "")
    return s


def load_targets_from_file(path) -> pd.DataFrame:
    """이미지 엑셀의 '이미지'가 아닌 시트(목표매출)에서 (공식/병행, 쇼핑몰, 월)별 목표 파싱.
    레이아웃: 'N월' 헤더 + 그 왼쪽 칸이 쇼핑몰명, 블록 라벨('병행'/'공식'), 합계행(쇼핑몰칸='쇼핑몰')."""
    cols = ["공식병행", "쇼핑몰", "월", "목표", "_key"]
    if path is None:
        return pd.DataFrame(columns=cols)
    try:
        xls = pd.ExcelFile(path)
    except Exception:
        return pd.DataFrame(columns=cols)
    sheets = [s for s in xls.sheet_names if str(s).strip() != "이미지"]
    if not sheets:
        return pd.DataFrame(columns=cols)
    try:
        raw = pd.read_excel(xls, sheet_name=sheets[0], header=None)
    except Exception:
        return pd.DataFrame(columns=cols)
    rows = raw.values.tolist()
    recs = []
    cur_block = None
    month_cols: dict[int, int] = {}
    for r in rows:
        cells = list(r)
        # 블록 헤더: 앞쪽 칸에 '병행'/'공식' + 행에 'N월' 라벨 존재
        blk = None
        for c in cells[:3]:
            cs = str(c).strip() if c is not None else ""
            if cs in ("병행", "공식"):
                blk = cs
                break
        has_month = any(c is not None and re.fullmatch(r"\d{1,2}월", str(c).strip()) for c in cells)
        if blk and has_month:
            cur_block = blk
            month_cols = {}
            for ci, c in enumerate(cells):
                m = re.fullmatch(r"(\d{1,2})월", str(c).strip()) if c is not None else None
                if m:
                    month_cols[ci] = int(m.group(1))
            continue
        if cur_block is None or not month_cols:
            continue
        name_idx = min(month_cols) - 1  # 쇼핑몰명은 월 컬럼 바로 왼쪽
        name = cells[name_idx] if 0 <= name_idx < len(cells) else None
        nm = str(name).strip() if name is not None else ""
        if (not nm) or nm.lower() == "nan" or nm in ("쇼핑몰", "목표매출"):
            continue  # 합계행/헤더/빈행
        for ci, mon in month_cols.items():
            v = pd.to_numeric(cells[ci], errors="coerce") if ci < len(cells) else np.nan
            if pd.notna(v) and v != 0:
                recs.append({"공식병행": cur_block, "쇼핑몰": nm, "월": int(mon),
                             "목표": float(v), "_key": _norm_mall(nm)})
    return pd.DataFrame(recs, columns=cols)


def find_image_file() -> Path | None:
    """앱 폴더에서 독립 이미지 매핑 파일('이미지*.xlsx' 등)을 찾는다(최신 우선)."""
    for pat in ("이미지*.xlsx", "이미지*.xls", "image*.xlsx", "images*.xlsx"):
        c = sorted(APP_DIR.glob(pat), key=lambda p: (p.stat().st_mtime, p.name), reverse=True)
        if c:
            return c[0]
    return None


def load_image_map_from_file(path) -> dict:
    """독립 '이미지' 엑셀(첫 시트): A열=라인명/모델명, B열=이미지URL → {키: url}."""
    if path is None:
        return {}
    try:
        df = pd.read_excel(path, sheet_name=0, header=None, usecols=[0, 1])
    except Exception:
        return {}
    mapping: dict[str, str] = {}
    for _, row in df.iterrows():
        name = row.iloc[0]
        url = str(row.iloc[1]).strip() if pd.notna(row.iloc[1]) else ""
        if pd.isna(name) or not url.lower().startswith(("http://", "https://")):
            continue
        for k in _img_keys(name):
            mapping[k] = url
    return mapping


@st.cache_data(show_spinner=False)
def load_image_map_from_image_xlsx(data: bytes | None) -> dict:
    """독립 '이미지' 엑셀에서 이미지 매핑을 읽는다. 시트명에 '이미지'가 있으면 그 시트,
    없으면 첫 시트의 A열=라인명/모델명, B열=이미지URL → {키: url}."""
    if data is None:
        return {}
    try:
        xls = pd.ExcelFile(io.BytesIO(data))
    except Exception:
        return {}
    target = None
    for s in xls.sheet_names:
        if "이미지" in str(s) or "image" in str(s).lower():
            target = s
            break
    if target is None:
        target = xls.sheet_names[0]
    try:
        df = pd.read_excel(xls, sheet_name=target, header=None, usecols=[0, 1])
    except Exception:
        return {}
    mapping: dict[str, str] = {}
    for _, row in df.iterrows():
        name = row.iloc[0]
        url = str(row.iloc[1]).strip() if pd.notna(row.iloc[1]) else ""
        if pd.isna(name) or not url.lower().startswith(("http://", "https://")):
            continue
        for k in _img_keys(name):
            mapping[k] = url
    return mapping


def find_stock_file() -> Path | None:
    """앱 폴더에서 재고 파일('재고*.xlsx','event_price*.xlsx' 등)을 찾는다(최신 우선)."""
    for pat in ("재고*.xlsx", "event_price*.xlsx", "재고*.xls", "stock*.xlsx"):
        c = sorted(APP_DIR.glob(pat), key=lambda p: (p.stat().st_mtime, p.name), reverse=True)
        if c:
            return c[0]
    return None


@st.cache_data(show_spinner=False)
def load_line_map(src) -> dict:
    """엑셀에서 모델명→라인명 매핑을 읽는다(라인명 '전용' 시트가 있을 때만).
    시트명에 '라인'/'line'/'매핑'이 들어간 시트만 사용하고, 헤더에서 '모델명'·'라인명' 열을
    찾아 매핑한다(열 순서 무관). 헤더가 없으면 A열=모델명, B열=라인명으로 본다.
    ※ '재고'/'목표' 같은 시트는 여기서 읽지 않는다(재고는 load_stock 이 따로 처리)."""
    if src is None:
        return {}
    try:
        xls = pd.ExcelFile(io.BytesIO(src)) if isinstance(src, (bytes, bytearray)) else pd.ExcelFile(src)
    except Exception:
        return {}
    target = None
    for s in xls.sheet_names:
        low = str(s).lower().strip()
        if any(k in low for k in ("라인", "line", "매핑")):
            target = s
            break
    if target is None:
        return {}
    try:
        raw = pd.read_excel(xls, sheet_name=target, header=None)
    except Exception:
        return {}
    rows = [list(r) for r in raw.where(pd.notna(raw), None).values.tolist()]
    if not rows:
        return {}
    mcol = lcol = None
    hdr_row = None
    for i in range(min(5, len(rows))):
        for j, c in enumerate(rows[i]):
            cs = str(c).replace(" ", "").replace("\n", "") if c is not None else ""
            if cs in ("모델명", "모델", "상품코드", "상품명") and mcol is None:
                mcol, hdr_row = j, i
            if cs in ("라인명", "라인") and lcol is None:
                lcol = j
        if mcol is not None and lcol is not None:
            break
    m: dict[str, str] = {}
    if mcol is not None and lcol is not None:
        for r in rows[(hdr_row or 0) + 1:]:
            if mcol < len(r) and lcol < len(r):
                a, b = r[mcol], r[lcol]
                ka = _norm_model(a)
                vb = str(b).strip() if b is not None else ""
                if ka and vb and ka.lower() not in ("nan", "none") and vb.lower() not in ("nan", "none"):
                    m[ka] = vb
    else:  # 헤더 못 찾음 → A열=모델명, B열=라인명 가정
        for r in rows:
            if len(r) >= 2:
                a, b = r[0], r[1]
                ka = _norm_model(a)
                vb = str(b).strip() if b is not None else ""
                if not ka or ka in ("모델명", "모델", "상품명") or vb in ("라인명", "라인", ""):
                    continue
                if ka.lower() in ("nan", "none") or vb.lower() in ("nan", "none"):
                    continue
                m[ka] = vb
    return m


_STOCK_HDR_KEYS = ("라인명", "브랜드", "모델명", "수량", "총원가", "원가평균", "가용수량")


def _xlsx_rows_fast_sheet(data: bytes, sheet_name) -> list:
    """openpyxl read_only 로 특정 시트만 값으로 빠르게 읽는다(대용량 재고 시트용)."""
    from openpyxl import load_workbook
    wb = load_workbook(io.BytesIO(data), read_only=True, data_only=True)
    try:
        ws = wb[sheet_name]
        rows = [list(r) for r in ws.iter_rows(values_only=True)]
    finally:
        wb.close()
    while rows and all(c is None for c in rows[-1]):
        rows.pop()
    return rows


_STOCK_CAT_DEBUG = {}  # 재고가 분류로 읽은 컬럼/매핑률 진단 기록


def _max_days_ago(text):
    """문자열의 'N일전' 들 중 가장 큰 N(가장 오래된 입고경과일) 반환. 여러 입고건/사이즈 대비. 없으면 NaN."""
    nums = re.findall(r"(\d+)\s*일\s*전", str(text))
    return max(int(n) for n in nums) if nums else np.nan


def _inbound_qty(text):
    """'N일전/M' 또는 'N일전M개' 들의 M(입고수량) 합 = 누적 입고량. 없으면 0."""
    pairs = re.findall(r"\d+\s*일\s*전\s*/?\s*([\d,]+)", str(text))
    return sum(int(m.replace(",", "")) for m in pairs) if pairs else 0


def _parse_inbound_events(text):
    """'N일전/M'(또는 'N일전M개') → [(N=경과일, M=수량), ...]. FIFO 입고 큐 구성용."""
    return [(int(n), int(m.replace(",", "")))
            for n, m in re.findall(r"(\d+)\s*일\s*전\s*/?\s*([\d,]+)", str(text))]


def _parse_inbound_cost(text):
    """S열 '날짜\\n수량 / 개당원가 비고' 들 → [(입고일(Timestamp), 개당원가), ...].
    각 입고건 = 'YYYY-MM-DD' + 그 뒤 첫 '/ 개당원가'. 시즌별 입고원가 산출용."""
    res = []
    for m in re.finditer(r"(\d{4}-\d{1,2}-\d{1,2})[\s\S]*?/\s*([\d,]+)", str(text)):
        try:
            res.append((pd.Timestamp(m.group(1)), int(m.group(2).replace(",", ""))))
        except Exception:
            continue
    return res


def _stock_rows_to_df(rows: list) -> pd.DataFrame:
    """재고 행 리스트 → 표준 재고 DataFrame[라인명,브랜드,모델명,수량,가용수량,원가평균,총원가].
    제목/숫자 행이 위에 있어도 헤더(라인명·브랜드·총원가…)를 자동 탐지한다."""
    if not rows:
        return pd.DataFrame()
    hdr_i = 0
    for i in range(min(8, len(rows))):
        vals = set(str(x).replace("\n", " ").strip() for x in rows[i] if x is not None)
        hit = sum(1 for k in _STOCK_HDR_KEYS if any(k == v or k in v for v in vals))
        if hit >= 4:
            hdr_i = i
            break
    header = _dedupe_cols([str(c).replace("\n", " ").strip() if c is not None else "" for c in rows[hdr_i]])
    sdf = pd.DataFrame(rows[hdr_i + 1:], columns=header)

    def col(*names):
        for name in names:  # 정확 매칭 우선
            key = name.replace(" ", "").replace("\n", "")
            for c in sdf.columns:
                if str(c).replace(" ", "").replace("\n", "") == key:
                    return c
        for name in names:  # 부분 매칭
            key = name.replace(" ", "").replace("\n", "")
            for c in sdf.columns:
                if key in str(c).replace(" ", "").replace("\n", ""):
                    return c
        return None

    c_line, c_brand, c_model = col("라인명"), col("브랜드"), col("모델명")
    c_qty, c_avail = col("수량"), col("가용수량")
    c_cost, c_total = col("원가평균"), col("총원가")
    c_daecat, c_cat = col("대카테고리"), col("카테고리")
    out = pd.DataFrame()
    out["라인명"] = sdf[c_line].astype(str).str.strip() if c_line else ""
    out["브랜드"] = sdf[c_brand].astype(str).str.strip() if c_brand else ""
    out["모델명"] = sdf[c_model].astype(str).str.strip() if c_model else out["라인명"]
    out["대카테고리"] = sdf[c_daecat].astype(str).str.strip() if c_daecat else ""
    out["카테고리"] = sdf[c_cat].astype(str).str.strip() if c_cat else ""
    # 대분류: 매출(_finalize_df)과 동일 규칙 — '대분류' 컬럼 있으면 그걸, 없으면 세부 '카테고리'에
    #   _CATEGORY_MAP 을 태워 8개 대분류로. 8개 밖이면 '미분류'.
    # 분류 컬럼 자동선택: 후보(대분류/중분류/카테고리/분류/품목 등) 중 8개로 가장 잘 매핑되는 컬럼.
    #   '대카테고리'(브랜드패션 류)는 제외. 매출처럼 '상의' 수준 컬럼이 있으면 그게 매핑률 1위로 잡힌다.
    _cands = []
    for _c in sdf.columns:
        _cn = str(_c).replace(" ", "").replace("\n", "")
        if "대카테고리" in _cn:
            continue
        if any(_k in _cn for _k in ["대분류", "중분류", "소분류", "카테고리", "분류", "품목", "구분", "종류"]):
            _cands.append(_c)
    _best_c, _best_r = None, -1.0
    _cand_report = {}
    for _c in _cands:
        _mp = sdf[_c].astype(str).str.strip().map(_classify_to8)
        _r = float((_mp != "미분류").mean()) if len(_mp) else 0.0
        _cand_report[str(_c)] = round(_r, 3)
        if _r > _best_r:
            _best_r, _best_c = _r, _c
    if _best_c is not None and _best_r > 0:
        _raw_cat = sdf[_best_c].astype(str).str.strip()
    else:
        _raw_cat = out["카테고리"]
    out["대분류"] = _raw_cat.map(_classify_to8)
    out["대분류_원본"] = _raw_cat.values
    _STOCK_CAT_DEBUG.clear()
    _STOCK_CAT_DEBUG.update({"selected": str(_best_c), "rate": round(_best_r, 3),
                             "candidates": _cand_report, "all_cols": [str(c) for c in sdf.columns]})
    out["수량"] = to_number(sdf[c_qty]) if c_qty else 0
    out["가용수량"] = to_number(sdf[c_avail]) if c_avail else np.nan
    out["원가평균"] = to_number(sdf[c_cost]) if c_cost else np.nan
    out["총원가"] = to_number(sdf[c_total]) if c_total else (out["수량"] * out["원가평균"])
    # 입고경과일(일수): I열(엑셀 9번째)에 'N일전' 형태. 한 셀에 여러 입고건이면 가장 오래된(=일수 최대).
    #   I열에 'N일전'이 없으면 입고 관련 헤더 컬럼으로 폴백.
    _hist = None
    if sdf.shape[1] >= 9:
        _cand_hist = sdf.iloc[:, 8]
        if _cand_hist.astype(str).str.contains(r"\d+\s*일\s*전", regex=True).mean() >= 0.3:
            _hist = _cand_hist
    if _hist is None:
        _hc = col("입고경과", "경과일", "입고이력", "입고내역", "입고")
        if _hc is not None:
            _hist = sdf[_hc]
    out["입고경과일행"] = _hist.astype(str).map(_max_days_ago) if _hist is not None else np.nan
    out["입고수량합행"] = _hist.astype(str).map(_inbound_qty) if _hist is not None else 0
    out["입고이벤트"] = _hist.astype(str).map(_parse_inbound_events) if _hist is not None else [[] for _ in range(len(out))]
    # S열(엑셀 19번째): '날짜 / 개당원가' 입고이력 → [(입고일, 개당원가)]. (I열 수량과 입고건 매칭용)
    _scost = None
    if sdf.shape[1] >= 19:
        _cand_s = sdf.iloc[:, 18]
        if _cand_s.astype(str).str.contains(r"\d{4}-\d{1,2}-\d{1,2}", regex=True).mean() >= 0.3:
            _scost = _cand_s
    out["입고원가이벤트"] = _scost.astype(str).map(_parse_inbound_cost) if _scost is not None else [[] for _ in range(len(out))]
    # 입고일자 보존 (시즌 SS/FW용). '입고이력' 처럼 날짜+수량+가격이 한 셀에 뭉친 형태도 처리:
    #   먼저 통째 파싱 → 비면 문자열에서 'YYYY-MM-DD' 추출. '입고예정'(전부 0) 같은 건 자동 제외.
    def _col_to_indate(s):
        d = _parse_date_flexible(s)
        if d.notna().mean() < 0.3:
            d2 = pd.to_datetime(
                s.astype(str).str.extract(r"(\d{4}-\d{1,2}-\d{1,2})", expand=False),
                errors="coerce")
            if d2.notna().mean() > d.notna().mean():
                d = d2
        return d

    _best_in = None
    c_in = col("입고일자", "입고일", "입고날짜", "입고일시", "입고이력")
    if c_in is not None:
        _best_in = _col_to_indate(sdf[c_in])
    if _best_in is None or _best_in.notna().mean() < 0.3:
        for _ic in sdf.columns:  # '입고' 포함 컬럼 중 날짜 추출률 최고를 채택
            if "입고" in str(_ic).replace(" ", ""):
                _cand = _col_to_indate(sdf[_ic])
                if _best_in is None or _cand.notna().mean() > _best_in.notna().mean():
                    _best_in = _cand
    out["입고일자"] = _best_in.values if _best_in is not None else pd.NaT
    out["공식/병행"] = [
        _classify_official(b, d, c) for b, d, c in zip(out["브랜드"], out["대카테고리"], out["카테고리"])
    ]
    out = out[~(out["라인명"].isin(["", "nan", "None"]) & out["모델명"].isin(["", "nan", "None"]))]
    return out.reset_index(drop=True)


@st.cache_data(show_spinner=False)
def load_stock(src) -> pd.DataFrame:
    """행사가관리/재고 엑셀(첫 시트) → 표준 재고 DataFrame."""
    if src is None:
        return pd.DataFrame()
    try:
        data = bytes(src) if isinstance(src, (bytes, bytearray)) else Path(src).read_bytes()
    except Exception:
        return pd.DataFrame()
    try:
        rows = _xlsx_rows_fast(data)
    except Exception:
        try:
            raw = pd.read_excel(io.BytesIO(data), header=None)
            rows = [list(r) for r in raw.where(pd.notna(raw), None).values.tolist()]
        except Exception:
            return pd.DataFrame()
    return _stock_rows_to_df(rows)


@st.cache_data(show_spinner=False)
def load_stock_from_image_xlsx(src) -> pd.DataFrame:
    """이미지 엑셀에 '재고'/'stock' 시트가 있으면 그 시트로 재고를 읽는다
    (별도 재고 파일을 올리지 않아도 됨)."""
    if src is None:
        return pd.DataFrame()
    try:
        data = bytes(src) if isinstance(src, (bytes, bytearray)) else Path(src).read_bytes()
    except Exception:
        return pd.DataFrame()
    try:
        from openpyxl import load_workbook
        wb = load_workbook(io.BytesIO(data), read_only=True)
        names = list(wb.sheetnames)
        wb.close()
    except Exception:
        return pd.DataFrame()
    target = None
    for s in names:
        if any(k in str(s).lower() for k in ("재고", "stock", "event_price", "행사가")):
            target = s
            break
    if target is None:
        return pd.DataFrame()
    try:
        rows = _xlsx_rows_fast_sheet(data, target)
    except Exception:
        return pd.DataFrame()
    return _stock_rows_to_df(rows)


def line_map_from_stock(stock_df: pd.DataFrame) -> dict:
    """재고 DataFrame 에서 모델명→라인명 매핑 추출(이미지 엑셀 매핑 보조)."""
    if stock_df is None or stock_df.empty:
        return {}
    m: dict[str, str] = {}
    for md, ln in zip(stock_df["모델명"], stock_df["라인명"]):
        k, v = _norm_model(md), str(ln).strip()
        if k and v and k.lower() not in ("nan", "none", "") and v.lower() not in ("nan", "none", ""):
            m.setdefault(k, v)
    return m


# ============================================================
# 원본(raw) 업로드 자동 전처리  ── process_excel.py 로직 포팅(값 전용)
#   HTML 위장 xls / 진짜 xlsx·xls 자동 인식 → 사은품 배송비 이전,
#   행 삭제, 쇼핑몰명 통일, 공식/병행 분류, 정산금·대분류 추가.
#   이미 가공된(완성) 파일이면 그대로 통과.
# ============================================================
# 열 인덱스(0-based) — process_excel.py 와 동일
_CI, _DI, _EI, _FI, _GI, _HI, _WI = 2, 3, 4, 5, 6, 8, 23
_CATEGORY_IDX, _M_IDX, _O_IDX = 7, 12, 14
_P_IDX, _Q_IDX, _ORDER_IDX, _BRAND_IDX = 15, 16, 1, 5
_TRUNCATE_IDX = 25
_NUM_COL_RANGE = range(9, 23)
_GIFT_BRANDS = ["쇼핑백", "사은품"]
_DELETE_H_VALUES = {"파슬AS", "쿠팡그로스 재고손실보상", "쿠팡그로스 기타정산"}
_DELETE_D_KEYWORDS = ["방송", "홈방", "나린인터", "태그바이"]
_D_RENAME_MAP = {
    "KREAM": "크림 주식회사", "카카오톡선물하기_디젤": "카카오톡선물하기",
    "카카오톡선물하기_병행": "카카오톡선물하기", "카카오톡선물하기_공식": "카카오톡선물하기",
    "에이블리(블리블리)": "에이블리", "에이블리(치페)": "에이블리",
    "무신사_블리블리": "무신사", "Wconcept(뷰티)": "Wconcept",
    "29CM(티켓투더문)": "29CM(공식)", "29CM(디젤)": "29CM(공식)",
    "카카오스타일 (치페)": "카카오스타일 (지그재그)",
    "카카오스타일 (티켓투더문)": "카카오스타일 (지그재그)",
    "카카오스타일 (블리블리)": "카카오스타일 (지그재그)",
}
_OFFICIAL_F_ONLY = {
    "블리블리", "헤브블루", "미스그린", "치페", "파슬", "아르마니", "티켓투더문",
    "아르마니익스체인지", "울프1834", "인도솔", "썬젤리", "스카겐", "미니쿄모", "스케쳐스",
}
_CATEGORY_MAP = {
    "가방": "가방", "귀걸이": "주얼리", "드레스": "의류", "라이터": "용품", "마사지볼": "용품",
    "모자": "소품", "목걸이": "주얼리", "문구": "용품", "반지": "주얼리", "밴드": "시계",
    "벨트": "소품", "상의": "의류", "시계": "시계", "신발": "신발", "아우터": "의류",
    "잡화ACC": "소품", "지갑": "지갑", "침낭": "용품", "키링&키홀더": "소품", "팔찌": "주얼리",
    "폼롤러": "용품", "하의": "의류", "핸드폰케이스": "소품", "홈데코": "용품", "우산": "소품",
    "옷걸이": "용품", "에어팟케이스": "용품", "언더웨어": "의류", "바디케어": "용품",
    "쇼핑백": "용품", "향수": "용품", "스킨케어": "용품", "거치대": "시계", "인솔": "용품",
    "쥬얼리보관함": "주얼리", "와인더": "시계", "시계보관함": "시계", "완구": "용품",
    "손난로": "용품", "참": "주얼리", "보온주머니": "용품", "생활잡화": "용품",
    "스포츠용품": "용품", "스윔웨어": "용품", "수납용품": "용품", "브로치": "소품",
    "케이블": "시계", "생활용품": "용품", "욕실용품": "용품", "슬립웨어": "의류",
    "아이메이크업": "용품", "립메이크업": "용품", "베이스메이크업": "용품", "뷰티소품": "용품",
    "클렌징": "용품", "선케어": "용품", "헤어케어": "용품", "주방용품": "용품",
}

# 대분류는 이 8개만 허용. _CATEGORY_MAP 으로 매핑 후 이 밖이면 '미분류'.
_ALLOWED_CATS = {"시계", "주얼리", "가방", "지갑", "의류", "신발", "소품", "용품"}

# 정확매칭(_CATEGORY_MAP)으로 안 잡히는 세부 명칭(반팔티/스니커즈 등)용 키워드 규칙.
#   더 구체적인 대분류를 앞에 둬 충돌 최소화(용품은 광범위해서 맨 뒤). 카테고리 글자에 키워드 포함 시 매칭.
_CAT_KEYWORDS = [
    ("시계", ["시계", "손목시계", "워치", "와인더"]),
    ("주얼리", ["목걸이", "귀걸이", "귀고리", "반지", "팔찌", "발찌", "브로치", "펜던트",
              "이어링", "이어커프", "네크리스", "뱅글", "앵클릿", "주얼리", "쥬얼리"]),
    ("지갑", ["지갑", "장지갑", "반지갑", "카드지갑", "카드케이스", "머니클립", "월렛", "코인케이스", "카드홀더"]),
    ("가방", ["가방", "백팩", "토트", "크로스", "숄더", "클러치", "파우치", "더플",
             "에코백", "메신저", "힙색", "보스턴", "버킷백", "호보", "쇼퍼백"]),
    ("신발", ["신발", "슈즈", "스니커", "운동화", "로퍼", "구두", "부츠", "샌들", "슬리퍼",
             "슬라이드", "뮬", "모카신", "더비", "옥스포드", "워커", "펌프스", "힐"]),
    ("소품", ["모자", "볼캡", "비니", "버킷햇", "페도라", "벨트", "스카프", "머플러", "목도리",
             "장갑", "양말", "삭스", "넥타이", "보타이", "손수건", "행커치프", "헤어밴드", "헤어핀",
             "머리끈", "머리띠", "집게핀", "선글라스", "안경테", "아이웨어", "키링", "키홀더"]),
    ("의류", ["반팔", "긴팔", "민소매", "나시", "티셔츠", "셔츠", "남방", "블라우스", "맨투맨",
             "후드", "후디", "니트", "스웨터", "가디건", "베스트", "조끼", "원피스", "드레스",
             "점프수트", "스커트", "치마", "팬츠", "바지", "슬랙스", "데님", "청바지", "조거",
             "트레이닝", "레깅스", "쇼츠", "반바지", "코트", "패딩", "다운", "점퍼", "자켓",
             "재킷", "블루종", "아노락", "아우터", "상의", "하의", "정장", "수트", "셋업",
             "폴로", "크롭", "슬립", "잠옷", "파자마", "언더웨어", "속옷", "브라", "드로즈",
             "러닝", "내의", "수영복", "스윔", "래쉬가드", "카라티"]),
    ("용품", ["향수", "퍼퓸", "코롱", "디퓨저", "캔들", "향초", "바디", "핸드크림", "풋크림",
             "로션", "크림", "세럼", "앰플", "에센스", "토너", "스킨", "클렌징", "마스크팩",
             "립밤", "마스카라", "섀도", "립스틱", "틴트", "쿠션팩트", "파운데이션", "컨실러",
             "선크림", "선스틱", "미스트", "헤어오일", "샴푸", "트리트먼트", "욕실", "주방",
             "텀블러", "보틀", "머그", "문구", "노트", "완구", "인형", "수납", "정리함",
             "폼롤러", "마사지", "우산", "양산", "담요", "블랭킷", "쿠션", "홈데코", "거치대",
             "충전기", "케이블", "보조배터리", "에어팟", "폰케이스", "그립톡", "스마트톡"]),
]


def _classify_to8(raw) -> str:
    """세부 카테고리를 8개 대분류로. 정확매칭(_CATEGORY_MAP) → 키워드 부분매칭 → 미분류."""
    s = str(raw).strip()
    if not s or s in ("nan", "None", "미분류"):
        return "미분류"
    v = _CATEGORY_MAP.get(s)
    if v in _ALLOWED_CATS:
        return v
    if s in _ALLOWED_CATS:          # 이미 대분류
        return s
    z = s.replace(" ", "")
    for cat, kws in _CAT_KEYWORDS:
        for kw in kws:
            if kw in z:
                return cat
    return "미분류"
_HDR_KEYWORDS = {"주문번호", "쇼핑몰", "브랜드", "수량", "최종판매가", "출고날짜"}
# 완성 파일의 컬럼명(위치 0~24). process_excel 이 위치(인덱스)로 처리하므로
# 원본 헤더명이 무엇이든 출력은 이 표준명을 위치 기준으로 박아 대시보드가 항상 인식하게 한다.
_CANON_HEADERS = [
    "차수", "주문번호", "품목코드", "쇼핑몰", "쇼핑몰아이디", "브랜드", "대카테고리", "카테고리",
    "모델명", "수량", "판매단가", "매출가", "최종판매가", "수수료", "수수료액", "마켓설정배송비",
    "실배송비", "출고원가", "원가총액", "수익원(마켓설정배송비)", "수익율(마켓설정배송비)",
    "수익원(실배송비)", "수익율(실배송비)", "출고날짜", "비고",
]


def _rawcell(row, idx):
    try:
        v = row[idx]
        return "" if v is None else str(v).strip()
    except (IndexError, KeyError):
        return ""


def _to_num(s):
    try:
        return float(str(s).replace(",", "").strip())
    except (ValueError, TypeError):
        return 0.0


def _get_d(row):
    d = _rawcell(row, _DI); e = _rawcell(row, _EI)
    try:
        if int(float(e)) == 1039456 and d.upper() == "GS SHOP":
            return "GS_API"
    except (ValueError, TypeError):
        pass
    if e == "033139LT":
        return "롯데홈쇼핑_API"
    return _D_RENAME_MAP.get(d, d)


def _classify_official(brand, daecat, cat) -> str:
    """브랜드 + 대카테고리 + 카테고리로 공식/병행 분류 (판매·재고 공통 로직)."""
    f = str(brand).strip(); g = str(daecat).strip(); h = str(cat).strip()
    if f in _OFFICIAL_F_ONLY:                               return "공식"
    if f == "마이클코어스" and g == "시계쥬얼리":          return "공식"
    if f == "디젤" and g == "시계쥬얼리":                  return "공식"
    if f == "라코스테" and g == "브랜드패션":              return "공식"
    if f == "토리버치" and h.startswith("TBW"):            return "공식"
    if f == "비비안웨스트우드" and h.startswith("VV"):     return "공식"
    return "병행"


def _get_c(row):
    return _classify_official(_rawcell(row, _FI), _rawcell(row, _GI), _rawcell(row, _HI))


def _should_delete(row):
    if _rawcell(row, _HI) in _DELETE_H_VALUES:
        return True
    return any(kw in _rawcell(row, _DI) for kw in _DELETE_D_KEYWORDS)


def _parse_raw_rows(data: bytes) -> list:
    """원본 파일 → 행 리스트(값). HTML 위장 / 진짜 xlsx 자동 인식. 무거운 파일도 빠르게."""
    head = data[:512].lower().lstrip()
    is_html = head[:1] == b"<" or b"<table" in head or b"<html" in head or b"<meta" in head
    if is_html:
        text = None
        for enc in ("utf-8", "cp949", "euc-kr"):
            try:
                text = data.decode(enc); break
            except Exception:
                text = None
        if text is None:
            text = data.decode("utf-8", errors="ignore")
        try:
            rows = _html_rows_fast(text)          # lxml 직접(빠름)
        except Exception:
            raw = pd.read_html(io.StringIO(text), header=None)[0]
            rows = [list(r) for r in raw.where(pd.notna(raw), None).values.tolist()]
    else:
        try:
            rows = _xlsx_rows_fast(data)          # openpyxl read_only(값만, 무거운 파일도 빠름)
        except Exception:
            raw = pd.read_excel(io.BytesIO(data), header=None, dtype=object)  # 구형 .xls(xlrd) 폴백
            rows = [list(r) for r in raw.where(pd.notna(raw), None).values.tolist()]
    return rows


def _is_already_processed(rows) -> bool:
    flat = set()
    for r in rows[:5]:
        for c in r:
            flat.add(str(c).strip() if c is not None else "")
    return "대분류" in flat or "정산금" in flat


def _find_header_idx(rows) -> int:
    """헤더(컬럼명) 행 인덱스. 없으면 -1(데이터가 0행부터 시작)."""
    for i in range(min(6, len(rows))):
        vals = set(str(x).strip() for x in rows[i] if x is not None)
        if len(vals & _HDR_KEYWORDS) >= 3:
            return i
    return -1


def _process_raw_rows(rows):
    """행 리스트 → (컬럼명, 데이터행) : 사은품 이전·삭제·분류·정산금/대분류."""
    hdr = _find_header_idx(rows)
    data_start = hdr + 1 if hdr >= 0 else 0  # 헤더 없으면 0행부터
    data_rows = [list(r) for r in rows[data_start:] if any(str(c).strip() for c in r if c is not None)]

    def is_gift(r):
        return _rawcell(r, _BRAND_IDX) in _GIFT_BRANDS

    # 사은품/쇼핑백 배송비를 같은 주문의 정상상품 첫 행으로 이전
    order_to_normal = {}
    for i, r in enumerate(data_rows):
        if not is_gift(r):
            order_to_normal.setdefault(_rawcell(r, _ORDER_IDX), []).append(i)
    orphan = set()
    for i, r in enumerate(data_rows):
        if not is_gift(r):
            continue
        targets = order_to_normal.get(_rawcell(r, _ORDER_IDX), [])
        if not targets:
            orphan.add(i); continue
        t = data_rows[targets[0]]
        for IDX in (_P_IDX, _Q_IDX):
            if IDX < len(r) and IDX < len(t):
                t[IDX] = _to_num(t[IDX]) + _to_num(r[IDX])
    kept = []
    for i, r in enumerate(data_rows):
        if is_gift(r) and i not in orphan:
            continue  # 매칭된 사은품 삭제
        if i in orphan and _BRAND_IDX < len(r):
            r[_BRAND_IDX] = _rawcell(r, _BRAND_IDX) + " ⚠미매칭"
        kept.append(r)
    data_rows = kept

    # 행 삭제 조건
    data_rows = [r for r in data_rows if not _should_delete(r)]

    # 쇼핑몰명 통일(D) / 공식·병행 분류(C)
    for r in data_rows:
        if _DI < len(r):
            r[_DI] = _get_d(r)
        if _CI < len(r):
            r[_CI] = _get_c(r)

    # 컬럼명: 원본 헤더명 대신 '표준명(위치 기준)' 사용 → 대시보드가 항상 인식
    col_names = list(_CANON_HEADERS[:_TRUNCATE_IDX])
    while len(col_names) < _TRUNCATE_IDX:
        col_names.append(f"col{len(col_names)}")
    col_names = col_names + ["정산금", "대분류"]

    # 데이터: 정산금(=최종판매가-수수료액)·대분류 추가
    out = []
    for r in data_rows:
        m = _to_num(r[_M_IDX]) if _M_IDX < len(r) else 0.0
        o = _to_num(r[_O_IDX]) if _O_IDX < len(r) else 0.0
        cat = _rawcell(r, _CATEGORY_IDX)
        대분류 = _CATEGORY_MAP.get(cat, "")
        r = list(r[:_TRUNCATE_IDX])
        while len(r) < _TRUNCATE_IDX:
            r.append("")
        r.append(round(m - o, 2))                       # 정산금
        r.append(대분류 if (대분류 or not cat) else "❓미매핑")  # 대분류
        out.append(r)
    return col_names, out


def _rows_to_xlsx_bytes(col_names, data_rows) -> bytes:
    """가공 결과를 완성 구조 xlsx 바이트로 (숫자/날짜 타입 지정)."""
    from openpyxl import Workbook
    import datetime as _dt
    wb = Workbook(); ws = wb.active
    ws.append(col_names)
    settle_idx = len(col_names) - 2  # 정산금 위치
    date_fmts = ["%Y-%m-%d", "%Y/%m/%d", "%Y%m%d", "%m/%d/%Y", "%d/%m/%Y"]
    for r in data_rows:
        row_out = []
        for ci, v in enumerate(r):
            if ci in _NUM_COL_RANGE or ci == settle_idx:
                row_out.append(_to_num(v))
            elif ci == _WI:
                if isinstance(v, (_dt.datetime, _dt.date, pd.Timestamp)):
                    row_out.append(pd.Timestamp(v).to_pydatetime())
                else:
                    s = str(v).strip()
                    s2 = s[:10] if (" " in s and len(s) >= 10) else s
                    dv = None
                    for fmt in date_fmts:
                        try:
                            dv = _dt.datetime.strptime(s2, fmt); break
                        except ValueError:
                            continue
                    row_out.append(dv if dv else s)
            else:
                row_out.append("" if v is None else v)
        ws.append(row_out)
    bio = io.BytesIO(); wb.save(bio)
    return bio.getvalue()


@st.cache_data(show_spinner=False)
def preprocess_upload(data: bytes) -> bytes:
    """원본이면 가공해 완성 구조 xlsx 바이트 반환. 이미 완성이면 원본 그대로."""
    try:
        rows = _parse_raw_rows(data)
    except Exception:
        return data  # 파싱 자체 실패 → 원본 그대로(로더가 직접 시도/에러)
    if not rows or _is_already_processed(rows):
        return data  # 이미 가공된(완성) 파일 → 그대로 통과
    col_names, out_rows = _process_raw_rows(rows)  # 실패 시 에러를 그대로 노출(원인 파악)
    return _rows_to_xlsx_bytes(col_names, out_rows)


# -----------------------------
# UI
# -----------------------------
st.title("🏷️ 브랜드 매출 대시보드")
st.caption("브랜드 1개 선택 → 시즌(SS/FW) 추이 · 쇼핑몰별 성과 · 카테고리/상품 비중")
st.info("왼쪽 사이드바에서 브랜드를 고르면 그 브랜드만 분석합니다. 3년+ 데이터면 SS/FW 시즌 비교가 핵심입니다.", icon="ℹ️")

with st.sidebar:
    st.header("데이터")
    PRINT_MODE = st.checkbox("📄 인쇄/PDF용 보기", value=False,
                             help="켜면 표가 정적 표로 바뀌어 인쇄·PDF가 깔끔하게 나옵니다. (정렬 기능은 꺼짐)")
    if PRINT_MODE:
        st.caption("인쇄 모드: 브라우저 인쇄(Ctrl+P) 후 끄세요.")
    uploaded = st.file_uploader(
        "Excel 파일 업로드 (여러 개 선택 가능)",
        type=["xlsx", "xls"],
        accept_multiple_files=True,
    )
    st.caption("판매 데이터 엑셀을 업로드하세요. (재고·이미지·목표는 같은 폴더의 '이미지.xlsx'·재고파일에서 자동 로드됩니다)")
    _imgf = find_image_file()
    if _imgf is not None:
        st.caption(f"🖼 이미지 매핑: **{_imgf.name}** (A열 라인명/모델명 · B열 이미지URL)")
    else:
        st.caption("🖼 상품 이미지: 같은 폴더에 '이미지.xlsx'(A열 라인명, B열 URL)를 두면 자동 표시됩니다.")
    st.divider()
    stock_up = st.file_uploader("📦 재고 파일 (선택)", type=["xlsx", "xls"], key="stock_upl")
    if stock_up is not None:
        st.caption("📦 재고 현황을 맨 아래에 표시합니다.")
    elif _imgf is not None:
        st.caption("📦 재고 현황: 이미지.xlsx 에 '재고' 시트가 있으면 자동으로 맨 아래에 표시됩니다. (없으면 재고 엑셀 업로드)")
    else:
        st.caption("📦 재고 현황: 행사가관리/재고 엑셀(라인명·브랜드·수량·총원가)을 올리면 표시됩니다.")

try:
    if uploaded:  # 1개 이상 업로드됨 (다중 허용 시 list)
        raws = [uf.getvalue() for uf in uploaded]
        frames = [load_upload(b) for b in raws]
        df = pd.concat(frames, ignore_index=True) if len(frames) > 1 else frames[0]
        img_map = {}
        for b in raws:  # 업로드 파일 안에 이미지 시트가 있으면 사용
            img_map.update(load_image_map_from_bytes(b))
        if len(uploaded) > 1:
            names = ", ".join(uf.name for uf in uploaded)
            st.success(f"📎 {len(uploaded)}개 파일 병합 분석 — {names} · 총 {len(df):,}행", icon="✅")
    else:
        st.info("왼쪽에서 **판매 데이터 엑셀**을 업로드하세요. "
                "재고·이미지·목표는 같은 폴더의 '이미지.xlsx'·재고파일에서 자동으로 불러옵니다.")
        st.stop()
    # 독립 '이미지' 엑셀 파일(앱 폴더의 이미지*.xlsx)이 있으면 병합 (우선 적용)
    _img_file = find_image_file()
    _img_bytes = _img_file.read_bytes() if _img_file is not None else None
    if _img_bytes is not None:
        img_map.update(load_image_map_from_image_xlsx(_img_bytes))
    targets_df = load_targets_from_file(_img_file)  # 목표매출(이미지 엑셀의 2번째 시트)
    # 모델명→라인명 매핑(라인명 전용 시트가 있을 때만) + 재고 로딩
    line_map = {}
    if _img_bytes is not None:
        line_map.update(load_line_map(_img_bytes))
    # 재고 소스 우선순위: 업로드 > 이미지.xlsx 의 '재고' 시트 > 앱 폴더 재고파일
    _stock_bytes = stock_up.getvalue() if stock_up else None
    stock_df = load_stock(_stock_bytes) if _stock_bytes else pd.DataFrame()
    if stock_df.empty and _img_bytes is not None:
        stock_df = load_stock_from_image_xlsx(_img_bytes)
    if stock_df.empty:
        _sf2 = find_stock_file()
        if _sf2 is not None:
            stock_df = load_stock(Path(_sf2).read_bytes())
    if not stock_df.empty:
        line_map.update(line_map_from_stock(stock_df))  # 재고의 모델명→라인명
    # 판매 데이터에 라인명 부여(베스트 상품을 라인명으로 취합)
    df["라인명"] = df["모델명"].apply(_line_of) if "모델명" in df.columns else ""
    # 재고 대분류 보강(매출 라인매핑) — 재고용 대분류 필터 옵션/필터에 사용
    if isinstance(stock_df, pd.DataFrame) and not stock_df.empty:
        if "대분류" not in stock_df.columns:
            stock_df["대분류"] = "미분류"
        stock_df["대분류"] = (stock_df["대분류"].astype(str).str.strip()
                            .replace({"": "미분류", "nan": "미분류", "None": "미분류"}))
        _sm = stock_df["대분류"] == "미분류"
        if _sm.any() and "라인명" in stock_df.columns:
            _l2c = (df.groupby("라인명")["대분류"]
                    .agg(lambda s: s.mode().iat[0] if len(s.mode()) else np.nan))
            _bf = stock_df.loc[_sm, "라인명"].astype(str).str.strip().map(_l2c)
            _bf = _bf.where(_bf.isin(_ALLOWED_CATS))
            stock_df.loc[_sm, "대분류"] = _bf.fillna("미분류").values
except Exception as e:
    st.error(f"파일을 읽는 중 오류가 발생했습니다: {e}")
    st.stop()


def make_product_display(prod_df: pd.DataFrame, extra_cols: list, img_width: str = "small"):
    """집계된 상품 표(모델명 포함)에 라인명→이미지 매칭하여 (표시용 df, column_config) 반환.
    img_map 이 비어있으면 이미지 컬럼 없이 그대로 표시. img_width: small/medium/large."""
    out = prod_df.copy().reset_index(drop=True)
    if "라인명" not in out.columns:
        out["라인명"] = out["모델명"].apply(to_line) if "모델명" in out.columns else ""
    show_img = bool(img_map)
    if show_img:
        out["이미지"] = out["라인명"].map(img_map).fillna("")
    out.insert(0, "Rank", np.arange(1, len(out) + 1))
    cols = ["Rank"] + (["이미지"] if show_img else []) + extra_cols
    disp, fcfg = format_table(out[cols])
    colcfg = {"Rank": st.column_config.TextColumn("#"), **fcfg}
    if show_img:
        colcfg["이미지"] = st.column_config.ImageColumn("이미지", width=img_width)
    return disp, colcfg


def product_cards_html(prod_df: pd.DataFrame, n: int = 10, img_px: int = 80, start: int = 1, step: int = 1) -> str:
    """상품을 카드형 HTML로 렌더 (이미지 크게, 순위는 인라인 #N).
    start/step: 순위 = start + i*step (좌우 교차 배치 시 step=2 등)."""
    rows = prod_df.head(n).reset_index(drop=True)
    show_imgs = bool(img_map)
    has_cat = "대분류" in rows.columns
    cards = []
    for i, r in rows.iterrows():
        rank = start + i * step
        brand = html.escape(str(r.get("브랜드", "")))
        cat = html.escape(str(r.get("대분류", ""))) if has_cat else ""
        season = ""
        if "시즌" in rows.columns:
            _sv = str(r.get("시즌", "")).strip()
            if _sv and _sv not in ("nan", "미상", "None"):
                season = html.escape(_sv)
        _parts = [f"#{rank}", brand] + ([cat] if cat else []) + ([season] if season else [])
        meta = " · ".join(_parts)
        line_name = str(r.get("라인명", "")).strip() or to_line(str(r.get("모델명", "")))
        model = html.escape(line_name if len(line_name) <= 38 else line_name[:37] + "…")
        rate = r.get("수익률", float("nan"))
        rate_s = f"{rate:.1f}%" if pd.notna(rate) and np.isfinite(rate) else "-"
        qty = r.get("수량", float("nan"))
        qty_s = f" · {int(qty):,}개" if pd.notna(qty) else ""
        sales_s = eok(r.get("최종판매가", 0))
        profit_s = eok(r.get("수익원(실배송비)", 0))
        img_block = ""
        if show_imgs:
            url = img_map.get(line_name, "")
            if url:
                img_block = (
                    f'<img src="{html.escape(url, quote=True)}" '
                    f'style="width:{img_px}px;height:{img_px}px;object-fit:cover;border-radius:8px;'
                    f'flex:0 0 auto;background:#f1f5f9;border:1px solid #eef2f7;">'
                )
            else:
                img_block = (
                    f'<div style="width:{img_px}px;height:{img_px}px;border-radius:8px;background:#f1f5f9;'
                    f'flex:0 0 auto;display:flex;align-items:center;justify-content:center;'
                    f'color:#cbd5e1;font-size:10px;">no img</div>'
                )
        cards.append(
            f'<div style="display:flex;gap:10px;align-items:center;padding:8px 8px;'
            f'border-bottom:1px solid #eef2f7;">{img_block}'
            f'<div style="min-width:0;flex:1;">'
            f'<div style="font-size:11px;color:#94a3b8;">{meta}</div>'
            f'<div style="font-size:13px;font-weight:600;color:#0f172a;white-space:nowrap;'
            f'overflow:hidden;text-overflow:ellipsis;">{model}</div>'
            f'<div style="font-size:13px;color:#0f172a;">{sales_s} '
            f'<span style="color:#64748b;">· 수익 {profit_s} · {rate_s}{qty_s}</span></div>'
            f'</div></div>'
        )
    return (
        '<div style="border:1px solid #e8eef5;border-radius:12px;overflow:hidden;'
        'box-shadow:0 2px 8px rgba(15,23,42,0.03);">' + "".join(cards) + "</div>"
    )


def stock_cards_html(stock_rows: pd.DataFrame, n: int = 10, img_px: int = 130, start: int = 1, step: int = 1) -> str:
    """재고 상품 카드형 HTML (판매 카드와 동일 레이아웃). 매출/수익률 대신 재고수량·총원가 표시."""
    rows = stock_rows.head(n).reset_index(drop=True)
    show_imgs = bool(img_map)
    has_cat = "대분류" in rows.columns
    cards = []
    for i, r in rows.iterrows():
        rank = start + i * step
        brand = html.escape(str(r.get("브랜드", "")))
        cat = html.escape(str(r.get("대분류", ""))) if has_cat else ""
        season = ""
        if "시즌" in rows.columns:
            _sv = str(r.get("시즌", "")).strip()
            if _sv and _sv not in ("nan", "미상", "None"):
                season = html.escape(_sv)
        _parts = [f"#{rank}", brand] + ([cat] if cat else []) + ([season] if season else [])
        meta = " · ".join(_parts)
        line_name = str(r.get("라인명", "")).strip() or to_line(str(r.get("모델명", "")))
        model = html.escape(line_name if len(line_name) <= 38 else line_name[:37] + "…")
        try:
            qty_i = int(float(r.get("재고수량", 0)))
        except Exception:
            qty_i = 0
        cost_s = eok(r.get("총원가", 0))
        el_s = ""
        _el = r.get("입고경과일")
        try:
            if _el is not None and not pd.isna(_el):
                el_s = f' · <span style="color:#b45309;">입고경과 {int(_el)}일</span>'
        except Exception:
            el_s = ""
        img_block = ""
        if show_imgs:
            url = img_map.get(line_name, "")
            if url:
                img_block = (
                    f'<img src="{html.escape(url, quote=True)}" '
                    f'style="width:{img_px}px;height:{img_px}px;object-fit:cover;border-radius:10px;'
                    f'flex:0 0 auto;background:#f1f5f9;border:1px solid #eef2f7;">'
                )
            else:
                img_block = (
                    f'<div style="width:{img_px}px;height:{img_px}px;border-radius:10px;background:#f1f5f9;'
                    f'flex:0 0 auto;display:flex;align-items:center;justify-content:center;'
                    f'color:#cbd5e1;font-size:11px;">no img</div>'
                )
        cards.append(
            f'<div style="display:flex;gap:12px;align-items:center;padding:10px 10px;'
            f'border-bottom:1px solid #eef2f7;">{img_block}'
            f'<div style="min-width:0;flex:1;">'
            f'<div style="font-size:11px;color:#94a3b8;">{meta}</div>'
            f'<div style="font-size:14px;font-weight:600;color:#0f172a;white-space:nowrap;'
            f'overflow:hidden;text-overflow:ellipsis;">{model}</div>'
            f'<div style="font-size:13px;color:#0f172a;">재고 {qty_i:,}개 '
            f'<span style="color:#64748b;">· 원가 {cost_s}</span>{el_s}</div>'
            f'</div></div>'
        )
    return (
        '<div style="border:1px solid #e8eef5;border-radius:12px;overflow:hidden;'
        'box-shadow:0 2px 8px rgba(15,23,42,0.03);">' + "".join(cards) + "</div>"
    )


def metric_cards_html(df: pd.DataFrame, value_fn, n: int = 10, img_px: int = 130,
                      start: int = 1, step: int = 1) -> str:
    """범용 카드(재고 카드와 동일 레이아웃). value_fn(row)이 값줄 HTML을 반환."""
    rows = df.head(n).reset_index(drop=True)
    show_imgs = bool(img_map)
    has_cat = "대분류" in rows.columns
    cards = []
    for i, r in rows.iterrows():
        rank = start + i * step
        brand = html.escape(str(r.get("브랜드", "")))
        cat = html.escape(str(r.get("대분류", ""))) if has_cat else ""
        season = ""
        if "시즌" in rows.columns:
            _sv = str(r.get("시즌", "")).strip()
            if _sv and _sv not in ("nan", "미상", "None"):
                season = html.escape(_sv)
        _parts = [f"#{rank}"] + ([brand] if brand else []) + ([cat] if cat else []) + ([season] if season else [])
        meta = " · ".join(_parts)
        line_name = str(r.get("라인명", "")).strip() or to_line(str(r.get("모델명", "")))
        model = html.escape(line_name if len(line_name) <= 38 else line_name[:37] + "…")
        img_block = ""
        if show_imgs:
            url = img_map.get(line_name, "")
            if url:
                img_block = (f'<img src="{html.escape(url, quote=True)}" '
                             f'style="width:{img_px}px;height:{img_px}px;object-fit:cover;border-radius:10px;'
                             f'flex:0 0 auto;background:#f1f5f9;border:1px solid #eef2f7;">')
            else:
                img_block = (f'<div style="width:{img_px}px;height:{img_px}px;border-radius:10px;background:#f1f5f9;'
                             f'flex:0 0 auto;display:flex;align-items:center;justify-content:center;'
                             f'color:#cbd5e1;font-size:11px;">no img</div>')
        cards.append(
            f'<div style="display:flex;gap:12px;align-items:center;padding:10px 10px;'
            f'border-bottom:1px solid #eef2f7;">{img_block}'
            f'<div style="min-width:0;flex:1;">'
            f'<div style="font-size:11px;color:#94a3b8;">{meta}</div>'
            f'<div style="font-size:14px;font-weight:600;color:#0f172a;white-space:nowrap;'
            f'overflow:hidden;text-overflow:ellipsis;">{model}</div>'
            f'{value_fn(r)}</div></div>'
        )
    return ('<div style="border:1px solid #e8eef5;border-radius:12px;overflow:hidden;'
            'box-shadow:0 2px 8px rgba(15,23,42,0.03);">' + "".join(cards) + "</div>")


VIBRANT_COLORS = [
    "#2563eb", "#f59e0b", "#10b981", "#ec4899", "#8b5cf6",
    "#06b6d4", "#ef4444", "#eab308", "#6366f1", "#14b8a6",
]


def share_donut(agg_df: pd.DataFrame, name_col: str, value_col: str, title: str, cmap: dict | None = None):
    """비중 도넛(생기있는 색). 슬라이스 라벨 = 이름 + 비중%만(크게). 값이 양수인 항목만."""
    d = agg_df.copy()
    d = d[pd.to_numeric(d[value_col], errors="coerce").fillna(0) > 0]
    names = d[name_col].astype(str).tolist()
    vals = list(d[value_col])
    if cmap:
        colors = [cmap.get(n, "#9aa5b1") for n in names]
    else:
        colors = [VIBRANT_COLORS[i % len(VIBRANT_COLORS)] for i in range(len(names))]
    fig = go.Figure(
        go.Pie(
            labels=names, values=vals, hole=0.5,
            textinfo="label+percent", textposition="inside",
            insidetextorientation="horizontal", textfont=dict(size=15),
            marker=dict(colors=colors, line=dict(color="#ffffff", width=2)),
            sort=False, direction="clockwise",
        )
    )
    fig.update_layout(title=title, showlegend=False, margin=dict(t=50, b=10, l=10, r=10),
                      font=dict(size=14))
    return fig


# =============================================================
# 브랜드 매출 대시보드 — 본문
#   브랜드 1개 선택 → 시즌(SS/FW) 추이 · 쇼핑몰별 · 카테고리/상품 비중
#   (공식/병행 구분 없음, 목표 없음)
# =============================================================

# ----- 시즌(SS/FW) 부여: '입고일자' 월 기준 -----
#   SS = 2~7월 입고, FW = 8~12월 입고, 1월 입고는 직전연도 FW 로 귀속.
#   관례가 다르면(예: SS=3~8월) 아래 SS_MONTHS 한 줄만 바꾸면 전체가 따라간다.
SS_MONTHS = {2, 3, 4, 5, 6, 7}


def _mk_one_season(d) -> str:
    """단일 날짜 → 시즌 라벨. NaT → '미상'. (_mk_season의 스칼라 버전, FIFO용)"""
    d = pd.Timestamp(d)
    if pd.isna(d):
        return "미상"
    m, y = d.month, int(d.year)
    if m in SS_MONTHS:
        return f"SS{y % 100:02d}"
    if m == 1:
        return f"FW{(y - 1) % 100:02d}"
    return f"FW{y % 100:02d}"


def _mk_season(dates) -> pd.Series:
    """datetime → 시즌 라벨('SS24'…). NaT → '미상'. SS_MONTHS 기준, 1월은 직전연도 FW."""
    d = pd.to_datetime(dates, errors="coerce")
    lab = pd.Series("미상", index=d.index, dtype=object)
    ok = d.notna()
    if ok.any():
        mm = d[ok].dt.month.values
        yy = d[ok].dt.year.values.astype(int)
        iss = np.isin(mm, list(SS_MONTHS))
        sy = np.where(iss, yy, np.where(mm == 1, yy - 1, yy)).astype(int)
        tp = np.where(iss, "SS", "FW")
        lab.loc[d.index[ok.values]] = [f"{t}{int(s) % 100:02d}" for t, s in zip(tp, sy)]
    return lab

# 시즌 기준 날짜 = '재고 파일의 입고일자'를 판매행에 매칭(모델명→라인명 순), 실패 시 출고날짜 폴백.
#   한 상품이 여러 번 입고됐으면 '최초 입고일'(min)을 시즌 기준으로 본다.
#   (최근 입고 기준으로 바꾸려면 아래 .min() 두 곳을 .max() 로)
def _instock_maps(sdf_stock):
    if not isinstance(sdf_stock, pd.DataFrame) or sdf_stock.empty or "입고일자" not in sdf_stock.columns:
        return {}, {}
    s = sdf_stock.copy()
    s["입고일자"] = _parse_date_flexible(s["입고일자"])
    s = s.dropna(subset=["입고일자"])
    # 오파싱(1970/2069 등) 방어: 합리적 연도만 사용
    s = s[s["입고일자"].dt.year.between(2010, 2035)]
    if s.empty:
        return {}, {}
    by_model = s.groupby(s["모델명"].astype(str).str.strip())["입고일자"].min()
    by_line = s.groupby(s["라인명"].astype(str).str.strip())["입고일자"].min()
    return by_model.to_dict(), by_line.to_dict()


def _events_by_model(sdf_stock):
    """재고 입고이벤트를 모델명별 [(입고일, 수량)] 리스트(입고일 오름차순)로. FIFO 큐용."""
    if (not isinstance(sdf_stock, pd.DataFrame) or sdf_stock.empty
            or "입고이벤트" not in sdf_stock.columns or "모델명" not in sdf_stock.columns):
        return {}
    today = pd.Timestamp.now().normalize()
    d = {}
    for m, evs in zip(sdf_stock["모델명"].astype(str).str.strip(), sdf_stock["입고이벤트"]):
        if not isinstance(evs, list) or not evs:
            continue
        bucket = d.setdefault(m, [])
        for n_days, qty in evs:
            try:
                bucket.append((today - pd.Timedelta(days=int(n_days)), int(qty)))
            except Exception:
                continue
    for m in d:
        d[m].sort(key=lambda x: x[0])  # 입고일 오름차순(오래된 것 먼저)
    return d


def _season_alloc_qty(ev_by_model, sold_by_model):
    """모델(사이즈)별 {시즌: 배분수량}. 출고일 무시, 오래된 입고시즌부터 판매량(net)을 순차로 채움.
    예: fw22 20개·ss24 2개 입고, 판매 21개 → {fw22:20, ss24:1}. 입고초과분은 최근 입고 시즌."""
    alloc = {}
    for model, evs in ev_by_model.items():
        sold = float(sold_by_model.get(model, 0.0))
        if sold <= 0:
            continue
        seq = sorted(evs, key=lambda x: x[0])  # 입고일 오름차순(오래된 것 먼저)
        aq, remain = {}, sold
        for dt, q in seq:
            if remain <= 0:
                break
            s = _mk_one_season(dt)
            take = min(remain, float(q))
            aq[s] = aq.get(s, 0.0) + take
            remain -= take
        if remain > 0:  # 입고이력보다 많이 팔림(과거 완판분 입고가 재고 스냅샷에 없음) → '미상'으로 분리
            aq["미상"] = aq.get("미상", 0.0) + remain
        if aq:
            alloc[model] = aq
    return alloc


def _split_df_by_season(sales_df, alloc, num_cols):
    """거래행을 모델별 시즌배분 비율로 분할(수치 안분). 시즌별 합계 = 순차배분값.
    같은 상품도 시즌이 다르면 별도 행으로 분리됨. 미배분(입고이력 없는) 모델은 '미상'."""
    if "모델명" not in sales_df.columns:
        return sales_df.assign(시즌="미상")
    keys = sales_df["모델명"].astype(str).str.strip()
    parts = []
    for model, grp in sales_df.groupby(keys):
        aq = alloc.get(model)
        tot = sum(aq.values()) if aq else 0.0
        if not aq or tot <= 0:
            g2 = grp.copy()
            g2["시즌"] = "미상"
            parts.append(g2)
            continue
        for s, q in aq.items():
            ratio = q / tot
            g2 = grp.copy()
            g2["시즌"] = s
            for c in num_cols:
                if c in g2.columns:
                    g2[c] = pd.to_numeric(g2[c], errors="coerce").fillna(0.0) * ratio
            parts.append(g2)
    return pd.concat(parts, ignore_index=True) if parts else sales_df.assign(시즌="미상")

_stock_for_season = stock_df if "stock_df" in dir() else pd.DataFrame()
_m_map, _l_map = _instock_maps(_stock_for_season)
_from_model = (df["모델명"].astype(str).str.strip().map(_m_map)
               if "모델명" in df.columns else pd.Series(pd.NaT, index=df.index))
_from_line = (df["라인명"].astype(str).str.strip().map(_l_map)
              if "라인명" in df.columns else pd.Series(pd.NaT, index=df.index))  # 시즌 추적용으로만 보존
# 매출 시즌 = 입고시즌별 순차배분(출고일 무시): 사이즈(모델)별 입고시즌 수량만큼 판매를 오래된 시즌부터 채우고,
#   같은 상품도 시즌이 다르면 거래행을 시즌별로 분리(수치 안분). 판매순위/회전율은 이후 라인명으로 합산.
_ev_by_model = _events_by_model(_stock_for_season)
_sold_by_model = (df.groupby(df["모델명"].astype(str).str.strip())["수량"].sum().to_dict()
                  if ("모델명" in df.columns and "수량" in df.columns) else {})
_alloc = _season_alloc_qty(_ev_by_model, _sold_by_model)
_num_cols_split = [c for c in ("수량", "최종판매가", "수익원(실배송비)", "목표", "원가", "원가총액") if c in df.columns]
df = _split_df_by_season(df, _alloc, _num_cols_split)
# 매칭률 = 시즌이 잡힌 매출 비중
_tot_sales = float(pd.to_numeric(df.get("최종판매가", pd.Series(dtype=float)), errors="coerce").fillna(0).sum())
_match_rate = (float(pd.to_numeric(df.loc[df["시즌"] != "미상", "최종판매가"], errors="coerce").fillna(0).sum()) / _tot_sales
               if _tot_sales else 0.0)
if not _ev_by_model:
    SEASON_SRC = "⚠ 재고 입고이력 없음 — 시즌 '미상' (재고 파일 / I열 'N일전/수량' 필요)"
    SEASON_FALLBACK = True
else:
    SEASON_SRC = f"입고시즌 순차배분 · 시즌 매칭 매출비중 {_match_rate:.0%} (미매칭은 미상)"
    SEASON_FALLBACK = False

# 시즌_타입/연도/정렬 = 시즌 라벨에서 역산
def _season_meta(lab):
    if not isinstance(lab, str) or lab == "미상" or len(lab) < 3:
        return ("미상", -1, 99999999)
    t = lab[:2]
    try:
        year = 2000 + int(lab[2:])
    except Exception:
        return ("미상", -1, 99999999)
    return (t, year, year * 10 + (1 if t == "SS" else 2))

_meta = df["시즌"].map(_season_meta)
df["시즌_타입"] = [x[0] for x in _meta]
df["시즌_연도"] = [x[1] for x in _meta]
df["시즌_정렬"] = [x[2] for x in _meta]

# 사은품/쇼핑백 제외 (매출 분석 공통 기준)
df = df[~df["브랜드"].astype(str).str.strip().isin(_GIFT_BRANDS)].copy()

season_order_all = (
    df[["시즌", "시즌_정렬"]].drop_duplicates().sort_values("시즌_정렬")["시즌"].tolist()
)
SEASON_MIN_YEAR = 2024  # 매출이 24년부터 시작 → 23년 이전 입고분은 모든 곳에서 제외


def _prev_same_season(s: str) -> str:
    """'SS24' → 전년 동시즌 'SS23' 라벨."""
    if not s or len(s) < 3:
        return ""
    typ, yy = s[:2], int(s[2:])
    return f"{typ}{(yy - 1) % 100:02d}"


# ----- 사이드바: 브랜드 선택 + 필터 -----
with st.sidebar:
    st.header("브랜드")
    _brand_tot = df.groupby("브랜드")["최종판매가"].sum().sort_values(ascending=False)
    brand_list = [b for b in _brand_tot.index.tolist()
                  if str(b).strip() not in ("미분류", "nan", "")]
    if not brand_list:
        st.error("브랜드 데이터가 없습니다.")
        st.stop()
    sel_brand = st.selectbox(
        "분석할 브랜드", brand_list, index=0,
        format_func=lambda b: f"{b}  ·  {eok(_brand_tot.get(b, 0))}",
    )
    st.caption(f"전체 {len(brand_list)}개 브랜드 · 선택 → **{sel_brand}**")
    st.divider()
    st.header("필터")
    bdf0 = df[df["브랜드"] == sel_brand].copy()

    _years_all = sorted(int(y) for y in bdf0["날짜"].dt.year.dropna().unique())
    sel_years = st.multiselect("연도 (판매·연도별 분석용)", _years_all,
                               default=_years_all[-3:] if _years_all else [])

    def _msa(label, options):
        opts = sorted([x for x in options if pd.notna(x)])
        return st.multiselect(label, opts, default=opts)

    sel_malls = _msa("쇼핑몰", bdf0["쇼핑몰"].unique())
    sel_cats = _msa("대분류 (판매)", bdf0["대분류"].unique())
    # 재고용 대분류 필터 (재고 섹션에만 적용)
    if isinstance(stock_df, pd.DataFrame) and not stock_df.empty and "대분류" in stock_df.columns:
        _bstk_opts = stock_df[stock_df["브랜드"].astype(str).str.strip()
                              == str(sel_brand).strip()]["대분류"].unique()
    else:
        _bstk_opts = []
    sel_cats_stock = _msa("대분류 (재고)", _bstk_opts) if len(_bstk_opts) else []
    include_returns = st.checkbox("반품/음수 포함", value=True)

# g: 브랜드 + (쇼핑몰/대분류/반품) 필터 — '시즌 전체' (시즌 추이·YoY 용)
g = bdf0[
    bdf0["쇼핑몰"].isin(sel_malls)
    & bdf0["대분류"].isin(sel_cats)
].copy()
if not include_returns:
    g = g[(g["수량"] >= 0) & (g["최종판매가"] >= 0)].copy()
if g.empty:
    st.warning("필터 조건에 해당하는 데이터가 없습니다.")
    st.stop()
# 23년 이전 입고분 제외 (매출이 24년부터 · 미상=입고불명은 유지)
g = g[(g["시즌_연도"] >= SEASON_MIN_YEAR) | (g["시즌_연도"] < 0)].copy()

# f: 판매 상세도 '전체 시즌' (시즌 필터 제거 — 판매량/매출은 라인 단위로 시즌 합산해서 봄)
f = g.copy()

season_order_g = [s for s in season_order_all
                  if s in set(g["시즌"].unique()) and s != "미상" and _season_meta(s)[1] >= SEASON_MIN_YEAR]
season_order = season_order_g
recent_seasons = season_order_g[-6:]  # 추이/구성 차트용 최근 6시즌

# 대분류 색맵 (도넛·시즌별구성 stacked 공유) — 이 브랜드 매출 비중 큰 순
_cat_order_all = g.groupby("대분류")["최종판매가"].sum().sort_values(ascending=False).index.tolist()
_CAT_CMAP = {c: VIBRANT_COLORS[i % len(VIBRANT_COLORS)] for i, c in enumerate(_cat_order_all)}

metric_cols = {
    "최종판매가": "최종판매가",
    "수량": "수량",
    "수익원(실배송비)": "수익원(실배송비)",
}

# ----- 브랜드 헤더 -----
st.markdown(
    f"<div style='font-size:1.6rem;font-weight:800;color:#0f172a;margin:.2rem 0 .2rem;'>🏷️ {html.escape(str(sel_brand))}</div>",
    unsafe_allow_html=True,
)

# ----- KPI -----
latest_season = season_order_g[-1] if season_order_g else None
prev_same = _prev_same_season(latest_season) if latest_season else ""
has_prev = bool(prev_same) and (prev_same in set(g["시즌"].unique()))

tot_sales = float(f["최종판매가"].sum())
tot_qty = float(f["수량"].sum())
tot_profit = float(f["수익원(실배송비)"].sum())
avg_price = tot_sales / tot_qty if tot_qty else 0
profit_rate = tot_profit / tot_sales * 100 if tot_sales else 0

ls_sales = float(g[g["시즌"] == latest_season]["최종판매가"].sum()) if latest_season else 0
ps_sales = float(g[g["시즌"] == prev_same]["최종판매가"].sum()) if has_prev else 0
season_yoy = (ls_sales - ps_sales) / abs(ps_sales) * 100 if ps_sales else np.nan

k = st.columns(5)
k[0].metric("총매출 (선택 시즌)", eok(tot_sales))
k[1].metric("수량", num(tot_qty))
k[2].metric("객단가", eok(avg_price))
k[3].metric("수익", eok(tot_profit), pct(profit_rate))
k[4].metric("수익률", pct(profit_rate))

st.markdown(
    f"<div class='hint'>선택 시즌 {len(season_order)}개 · {len(f):,}행 · "
    f"기간 {f['날짜'].min().date()} ~ {f['날짜'].max().date()} · 사은품/쇼핑백 제외</div>",
    unsafe_allow_html=True,
)

# =============================================================
# 1) 시즌(SS/FW) 추이
# =============================================================
st.markdown(f"<div class='section-title'>시즌(SS/FW) 추이 — {sel_brand}</div>", unsafe_allow_html=True)
st.caption(f"시즌 기준: **{SEASON_SRC}** · SS=2~7월 입고 · FW=8~12월 입고(1월 입고는 직전연도 FW). 입고시즌별 순차배분(출고일 무시·반품 차감)이며, 재고 입고이력이 없는 건 '미상'.")
if SEASON_FALLBACK:
    st.warning("재고 입고이력(I열 'N일전M개')이 없어(또는 재고 미업로드) 모든 매출 시즌이 '미상'입니다. 재고 파일을 올리고 입고이력 컬럼이 있는지 확인하세요.", icon="⚠️")
elif _match_rate < 0.7:
    st.warning(f"판매의 **{_match_rate:.0%}**만 재고 입고이력과 순차배분 매칭됐고 나머지는 '미상'입니다. 재고는 보통 '현재 보유분' 스냅샷이라 완판/단종 상품은 입고이력이 없어 과거일수록 미상이 많습니다. (정확히 하려면 전 기간 입고이력이 담긴 재고 파일 필요)", icon="⚠️")

with st.expander("🔧 시즌(입고일자) 진단 — 이상값 확인용"):
    st.write(f"FIFO 입고이력: 모델 **{len(_ev_by_model)}개** · 판매 시즌 매칭률 **{_match_rate:.0%}**")
    _sc = {k: int(v) for k, v in df["시즌"].value_counts().sort_index().items()}
    st.write("시즌별 행수:", _sc)
    if _m_map:
        _keys = list(_m_map.keys())[:15]
        _smp = pd.DataFrame({"모델명": _keys,
                             "입고일자(파싱결과)": [pd.Timestamp(_m_map[k]).date() for k in _keys]})
        render_table(_smp, hide_index=True, use_container_width=True)
    else:
        st.caption("재고에서 입고일자를 못 읽었습니다 (재고 미업로드 / 입고이력 컬럼 없음 / 형식 문제). 이 경우 매출 시즌은 '미상'입니다.")

    # ---- 특정 상품이 FIFO로 어떻게 시즌 나뉘는지 추적 ----
    st.markdown("---")
    _q = st.text_input("시즌 추적 — 모델명 입력 (예: 50CMA0069-55N-1Y5)", "", key="season_trace")
    if _q.strip():
        _k = _q.strip()
        _row = df[df["모델명"].astype(str).str.strip() == _k]
        _ev = _ev_by_model.get(_k)
        st.write("**입고이력 (오래된 입고순 · 시즌):**")
        if _ev:
            for _dt, _q2 in sorted(_ev, key=lambda x: x[0]):
                st.write(f"   · {pd.Timestamp(_dt).date()} ({_mk_one_season(_dt)}) — {int(_q2):,}개")
        else:
            st.write("   · 재고 입고이력 없음 → 이 모델 판매는 모두 '미상'")
        if not _row.empty:
            _vc = _row.groupby("시즌")["수량"].sum().sort_index()
            st.write("**이 모델 판매의 FIFO 시즌 분포:** "
                     + ", ".join(f"{s} {int(v):,}개" for s, v in _vc.items()))
            st.write(f"- 거래 {len(_row)}건 · 총 판매수량 {int(_row['수량'].sum()):,}개 "
                     f"(가장 오래된 입고시즌부터 순차 배분 · 출고일 무시)")
        else:
            st.caption("판매 데이터에서 이 모델명을 찾지 못했습니다.")

    # ---- 재고 입고이력 파싱 진단 (총입고원가 0 문제 확인용) ----
    st.markdown("---")
    st.write("**재고 입고이력 파싱 진단 (이 브랜드)**")
    _sdbg = (_stock_for_season[_stock_for_season["브랜드"].astype(str).str.strip() == str(sel_brand).strip()]
             if ("브랜드" in _stock_for_season.columns and not _stock_for_season.empty) else _stock_for_season)
    if isinstance(_sdbg, pd.DataFrame) and not _sdbg.empty:
        if "입고이벤트" in _sdbg.columns:
            st.write("I열 입고이벤트(수량) 샘플:", [e for e in _sdbg["입고이벤트"].head(5) if e][:5])
        if "입고원가이벤트" in _sdbg.columns:
            _cs = [e for e in _sdbg["입고원가이벤트"].head(20) if e][:5]
            st.write("S열 입고원가이벤트(원가) 샘플:", _cs if _cs else "⚠ 전부 비어있음 (S열 파싱 실패)")
        else:
            st.write("⚠ S열 입고원가이벤트 컬럼 자체가 없음")
        if isinstance(_stock_for_season, pd.DataFrame) and _stock_for_season.shape[1] >= 19:
            st.caption(f"참고 — 재고 19번째 컬럼(S열로 가정) 원본 샘플: "
                       f"{[str(x) for x in _stock_for_season.iloc[:5, 18].tolist()]}")
        st.caption(f"재고 전체 컬럼({_stock_for_season.shape[1]}개): "
                   f"{[str(c) for c in _stock_for_season.columns.tolist()]}")
    else:
        st.caption("이 브랜드 재고가 없습니다.")

sea = (g[g["시즌"] != "미상"].groupby("시즌")
         .agg(매출=("최종판매가", "sum"), 수량=("수량", "sum"),
              수익=("수익원(실배송비)", "sum"), 정렬=("시즌_정렬", "first"))
         .reset_index().sort_values("정렬"))
sea["타입"] = np.where(sea["시즌"].str.startswith("SS"), "SS", "FW")
sea["라벨"] = sea["매출"].apply(eok)

sc1, sc2 = st.columns([1.15, 1.5])
with sc1:
    fig_s = px.bar(
        sea, x="시즌", y="매출", color="타입", text="라벨",
        title=f"{sel_brand} 시즌별 매출",
        category_orders={"시즌": sea["시즌"].tolist(), "타입": ["SS", "FW"]},
        color_discrete_map={"SS": "#10b981", "FW": "#6366f1"},
        labels={"매출": "매출", "시즌": "시즌"},
    )
    fig_s.update_traces(textposition="outside", textangle=0, cliponaxis=False)
    fig_s.update_layout(xaxis_type="category", legend_title_text="시즌",
                        margin=dict(t=54, b=10), uniformtext_minsize=9, uniformtext_mode="hide")
    if len(sea):
        _ymax = float(sea["매출"].max())
        _ymin = float(sea["매출"].min())
        fig_s.update_yaxes(range=[min(0, _ymin) * 1.1, _ymax * 1.18 if _ymax > 0 else _ymax * 0.8])
    st.plotly_chart(fig_s, use_container_width=True)

with sc2:
    yr = (g[g["시즌_타입"] != "미상"].groupby(["시즌_연도", "시즌_타입"])["최종판매가"]
          .sum().unstack(fill_value=0))
    yr = yr[yr.index >= 0]
    for _t in ("SS", "FW"):
        if _t not in yr.columns:
            yr[_t] = 0.0
    yr = yr[["SS", "FW"]].sort_index()
    yr["합계"] = yr["SS"] + yr["FW"]
    _byr_prof = g[g["시즌_연도"] > 0].groupby("시즌_연도")["수익원(실배송비)"].sum()
    yt = pd.DataFrame({"연도": [f"{int(y)}" for y in yr.index]})
    yt["합계매출"] = yr["합계"].round(0).astype("int64").values
    yt["수익"] = [int(round(float(_byr_prof.get(y, 0.0)))) for y in yr.index]
    yt["수익률"] = [float(_byr_prof.get(y, 0.0)) / yr.loc[y, "합계"] * 100 if yr.loc[y, "합계"] else np.nan
                  for y in yr.index]
    yt["전년比"] = (yr["합계"].pct_change() * 100).apply(growth_pct).values
    yt["SS"] = yr["SS"].round(0).astype("int64").values
    yt["FW"] = yr["FW"].round(0).astype("int64").values
    st.markdown("**연도별 매출 · 수익**")
    _ycfg = {c: st.column_config.NumberColumn(c, format="localized") for c in ("합계매출", "수익", "SS", "FW")}
    _ycfg["수익률"] = st.column_config.NumberColumn("수익률", format="%.1f%%")
    render_table(yt, hide_index=True, use_container_width=True, height=60 + len(yt) * 36,
                 column_config=_ycfg)
    if has_prev:
        st.caption(f"최근 **{latest_season}** {eok(ls_sales)} · 전년 동시즌 {eok(ps_sales)} → **{growth_pct(season_yoy)}**")

# ----- 연도별 재고 소진 (시즌별) : 입고원가(I열 수량 × S열 원가) vs 현재 재고원가(V열 총원가) -----
_bsea = (_stock_for_season[_stock_for_season["브랜드"].astype(str).str.strip() == str(sel_brand).strip()]
         if ("브랜드" in _stock_for_season.columns) else _stock_for_season.iloc[0:0])
_in_cost, _cur_cost = {}, {}
if (not _bsea.empty) and ("입고이벤트" in _bsea.columns) and ("입고원가이벤트" in _bsea.columns):
    _t0b = pd.Timestamp.now().normalize()
    for _qev, _cev in zip(_bsea["입고이벤트"], _bsea["입고원가이벤트"]):
        if not isinstance(_qev, list):
            continue
        # S열: 시즌별 개당원가 (입고건 날짜 → 시즌)
        _c_by_s = {}
        for _d, _c in (_cev if isinstance(_cev, list) else []):
            _c_by_s[_mk_one_season(_d)] = float(_c)
        # I열: 시즌별 수량 × 그 시즌 개당원가
        for _n, _qq in _qev:
            _s = _mk_one_season(_t0b - pd.Timedelta(days=int(_n)))
            if _s == "미상" or _season_meta(_s)[1] < SEASON_MIN_YEAR:
                continue
            _c = _c_by_s.get(_s, 0.0)
            _in_cost[_s] = _in_cost.get(_s, 0.0) + float(_qq) * _c
if (not _bsea.empty) and ("총원가" in _bsea.columns) and ("입고일자" in _bsea.columns):
    _amt2 = pd.to_numeric(_bsea["총원가"], errors="coerce").fillna(0.0)
    _s2 = _mk_season(_bsea["입고일자"])
    for _a, _slab in zip(_amt2, _s2):
        if isinstance(_slab, str) and _slab != "미상" and _season_meta(_slab)[1] >= SEASON_MIN_YEAR:
            _cur_cost[_slab] = _cur_cost.get(_slab, 0.0) + float(_a)
_sea_prof = (g[g["시즌"] != "미상"].groupby("시즌")["수익원(실배송비)"].sum()
             if "수익원(실배송비)" in g.columns else pd.Series(dtype=float))
_cost_col = next((c for c in ("원가총액", "원가") if c in g.columns), None)
_sea_cost = (g[g["시즌"] != "미상"].groupby("시즌")[_cost_col].sum()
             if _cost_col else pd.Series(dtype=float))
_all_s = {s for s in (set(_in_cost) | set(_cur_cost) | set(season_order_g)) if _season_meta(s)[1] >= SEASON_MIN_YEAR}
_seasons_sorted = sorted(_all_s, key=lambda s: _season_meta(s)[2])
if _seasons_sorted:
    st2 = pd.DataFrame({"시즌": _seasons_sorted})
    st2["총입고원가"] = [int(round(_in_cost.get(s, 0.0))) for s in _seasons_sorted]
    st2["현재총원가"] = [int(round(_cur_cost.get(s, 0.0))) for s in _seasons_sorted]
    st2["소진율%"] = [min(max((_in_cost.get(s, 0) - _cur_cost.get(s, 0)) / _in_cost.get(s, 0) * 100, 0.0), 100.0)
                   if _in_cost.get(s, 0) else np.nan for s in _seasons_sorted]
    st2["원가회수율%"] = [(_sea_prof.get(s, 0.0) + _sea_cost.get(s, 0.0)) / _in_cost.get(s, 0.0) * 100
                      if _in_cost.get(s, 0.0) else np.nan for s in _seasons_sorted]
    st.markdown("**연도별 재고 소진** (시즌별 · 입고원가 vs 현재 재고원가)")
    _s2cfg = {c: st.column_config.NumberColumn(c, format="localized") for c in ("총입고원가", "현재총원가")}
    _s2cfg["소진율%"] = st.column_config.NumberColumn("소진율%", format="%.1f%%")
    _s2cfg["원가회수율%"] = st.column_config.NumberColumn("원가회수율%", format="%.1f%%")
    render_table(st2, hide_index=True, use_container_width=True, height=60 + len(st2) * 36,
                 column_config=_s2cfg)
    st.caption("총입고원가 = I열 입고수량 × S열 개당원가(시즌별) · 현재총원가 = V열(남은 재고 총원가) · "
               "소진율 = (총입고원가−현재총원가)÷총입고원가 · "
               "원가회수율 = (판매수익+판매원가)÷총입고원가 (입고원가 대비 판매로 회수한 금액).")
else:
    st.caption("※ 연도별 재고 소진표는 재고 파일(I열 입고수량 · S열 입고원가 · V열 총원가)이 있어야 표시됩니다.")

# =============================================================
# 1-2) 분기별 판매 추이 (판매=출고일 기준)
# =============================================================
st.markdown(f"<div class='section-title'>분기별 판매 추이 — {sel_brand}</div>", unsafe_allow_html=True)
st.caption("판매(출고)일 기준 · 분기별 매출. 시즌(입고일 기준)과 다른 축이라 시즌 필터와 무관하게 전체 판매기간을 봅니다.")

q = g.copy()
_qp = q["날짜"].dt.to_period("Q")
q["분기"] = _qp.astype(str)                       # 2024Q1
q["분기정렬"] = _qp.astype("int64")
q["분기표시"] = q["분기"].str.replace("Q", " Q", regex=False)
q["분기연도"] = q["날짜"].dt.year.astype(int)
q["분기No"] = q["날짜"].dt.quarter.astype(int)
qa = (q.groupby(["분기표시"])
        .agg(매출=("최종판매가", "sum"), 수량=("수량", "sum"),
             정렬=("분기정렬", "first"), 연도=("분기연도", "first"))
        .reset_index().sort_values("정렬"))
qa["라벨"] = qa["매출"].apply(eok)
qa["연도"] = qa["연도"].astype(int).astype(str)

qc1, qc2 = st.columns([1.7, 1])
with qc1:
    fig_q = px.bar(qa, x="분기표시", y="매출", text="라벨", color="연도",
                   title=f"{sel_brand} 분기별 매출(판매일 기준 · 연도별 색)",
                   labels={"매출": "매출", "분기표시": "분기"},
                   color_discrete_sequence=VIBRANT_COLORS)
    fig_q.update_traces(textposition="outside", textangle=0, cliponaxis=False)
    fig_q.update_layout(xaxis_type="category", margin=dict(t=54, b=10),
                        uniformtext_minsize=9, uniformtext_mode="hide", legend_title_text="연도")
    fig_q.update_xaxes(categoryorder="array", categoryarray=qa["분기표시"].tolist(),
                       tickangle=-30, title_text="")
    fig_q.update_yaxes(title_text="")
    if len(qa):
        _qm = float(qa["매출"].max()); _qn = float(qa["매출"].min())
        fig_q.update_yaxes(range=[min(0, _qn) * 1.1, _qm * 1.18 if _qm > 0 else _qm * 0.8])
    st.plotly_chart(fig_q, use_container_width=True)

with qc2:
    # 분기 × 연도 (전년 동분기 비교)
    qpv = q.pivot_table(index="분기No", columns="분기연도", values="최종판매가",
                        aggfunc="sum", fill_value=0).sort_index()
    if not qpv.empty:
        qd = pd.DataFrame({"분기": [f"Q{int(i)}" for i in qpv.index]})
        for _yy in qpv.columns:
            qd[str(int(_yy))] = qpv[_yy].round(0).astype("int64").values
        st.markdown("**분기 × 연도 매출**")
        _qcfg = {str(int(c)): st.column_config.NumberColumn(str(int(c)), format="localized")
                 for c in qpv.columns}
        render_table(qd, hide_index=True, use_container_width=True, height=60 + len(qd) * 36,
                     column_config=_qcfg)
        _yrs = sorted(int(c) for c in qpv.columns)
        if len(_yrs) >= 2:
            _a, _b = _yrs[-2], _yrs[-1]
            _ga = float(qpv[_a].sum()); _gb = float(qpv[_b].sum())
            _gr = (_gb - _ga) / abs(_ga) * 100 if _ga else np.nan
            st.caption(f"{_b}년 {eok(_gb)} · 전년 {eok(_ga)} → **{growth_pct(_gr)}**")

# =============================================================
# 2) 쇼핑몰별 — 이 브랜드가 어디서 잘 나가나
# =============================================================
st.markdown(f"<div class='section-title'>쇼핑몰별 — {sel_brand}</div>", unsafe_allow_html=True)

mc1, mc2 = st.columns([1, 1.3])
with mc1:
    mb = aggregate(f, ["쇼핑몰"], metric_cols).head(10).copy()
    if not mb.empty:
        mb["라벨"] = mb["매출비중"].apply(lambda x: f"{x:.1f}%")
        fig_m = px.bar(mb, x="쇼핑몰", y="최종판매가", text="라벨",
                       title=f"{sel_brand} 쇼핑몰 TOP 10 (매출 비중)",
                       labels={"최종판매가": "매출"})
        fig_m.update_traces(textposition="inside", insidetextanchor="middle", textangle=0,
                            textfont=dict(size=14, color="#ffffff"), marker_color="#2563eb",
                            cliponaxis=False)
        fig_m.update_layout(xaxis_type="category", margin=dict(t=54, b=10),
                            uniformtext_minsize=10, uniformtext_mode="hide")
        fig_m.update_xaxes(categoryorder="total descending", tickangle=-30, title_text="",
                           tickfont=dict(size=11))
        fig_m.update_yaxes(title_text="")
        st.plotly_chart(fig_m, use_container_width=True)
with mc2:
    mall_t = aggregate(f, ["쇼핑몰"], metric_cols).reset_index(drop=True)
    mall_t = mall_t[["쇼핑몰", "수량", "최종판매가", "객단가", "수익률", "매출비중"]].head(50)
    mall_t = rank_table(mall_t, "쇼핑몰")
    _ft, _fc = format_table(mall_t)
    render_table(_ft, hide_index=True, use_container_width=True, height=430,
                 column_config={**_fc, "Rank": st.column_config.TextColumn("#")})

# 연도별 실적 + 쇼핑몰 × 연도 (판매일 기준 · 사이드바 '연도' 필터 적용)
_gy = g.copy()
_gy["연도"] = _gy["날짜"].dt.year
_yrs = [int(y) for y in (sel_years or []) if y in set(_gy["연도"].dropna().astype(int))]
if not _yrs:
    _yrs = sorted(int(y) for y in _gy["연도"].dropna().unique())[-3:]
_gy = _gy[_gy["연도"].isin(_yrs)]
if not _gy.empty:
    st.markdown("**연도별 쇼핑몰 실적** (연도별 3분할 · 판매일 기준)")
    _T = float(_gy["최종판매가"].sum())
    _P = float(_gy["수익원(실배송비)"].sum())
    _Q = float(_gy["수량"].sum())
    _PR = (_P / _T * 100) if _T else 0.0
    st.markdown(f"합계({' · '.join(str(y) for y in _yrs)}) · "
                f"**총매출 {eok(_T)}** · 수익 {eok(_P)} · 수익률 {_PR:.1f}% · 수량 {int(_Q):,}개")
    _gall = g.copy()
    _gall["연도"] = _gall["날짜"].dt.year
    for _box, _yv in zip(st.columns(len(_yrs)), _yrs):
        with _box:
            _cur = _gall[_gall["연도"] == _yv]
            _prev = _gall[_gall["연도"] == _yv - 1]
            _tot = float(_cur["최종판매가"].sum())
            _prof = float(_cur["수익원(실배송비)"].sum())
            _pr = (_prof / _tot * 100) if _tot else 0.0
            _ptot = float(_prev["최종판매가"].sum())
            _yoy = ((_tot - _ptot) / _ptot * 100) if _ptot else np.nan
            st.markdown(f"**{_yv}년** · 총매출 {eok(_tot)}")
            st.caption(f"수익률 {_pr:.1f}% · 전년比 {growth_pct(_yoy)}")
            _a = _cur.groupby("쇼핑몰").agg(매출=("최종판매가", "sum"), 수량=("수량", "sum"),
                                          수익=("수익원(실배송비)", "sum"))
            if _a.empty:
                st.caption("데이터 없음")
                continue
            _a["객단가"] = np.where(_a["수량"] != 0, _a["매출"] / _a["수량"], 0)
            _a["수익률"] = np.where(_a["매출"] != 0, _a["수익"] / _a["매출"] * 100, 0)
            _pm = _prev.groupby("쇼핑몰")["최종판매가"].sum()
            _pv = pd.Series(_a.index.map(_pm), index=_a.index).astype(float)
            _a["전년비"] = np.where(_pv.notna() & (_pv != 0), (_a["매출"] - _pv) / _pv * 100, np.nan)
            _a = _a.sort_values("매출", ascending=False).head(20).reset_index()
            _d = pd.DataFrame({
                "쇼핑몰": [f"{i}. {v}" for i, v in enumerate(_a["쇼핑몰"], 1)],
                "매출": _a["매출"].round(0).astype("int64").values,
                "객단가": _a["객단가"].round(0).astype("int64").values,
                "수익률": _a["수익률"].round(1).values,
                "전년비": pd.Series(_a["전년비"]).apply(growth_pct).values,
            })
            render_table(_d, hide_index=True, use_container_width=True,
                         height=min(60 + len(_d) * 36, 620),
                         column_config={
                             "매출": st.column_config.NumberColumn("매출", format="localized"),
                             "객단가": st.column_config.NumberColumn("객단가", format="localized"),
                             "수익률": st.column_config.NumberColumn("수익률", format="%.1f%%")})

# =============================================================
# 3) 카테고리(대분류) 비중·매출
# =============================================================
st.markdown(f"<div class='section-title'>카테고리(대분류) · 매출 기준 — {sel_brand}</div>", unsafe_allow_html=True)
cat_t = aggregate(f, ["대분류"], metric_cols).reset_index(drop=True)
cc1, cc2 = st.columns([1.3, 1])
with cc1:
    st.plotly_chart(share_donut(cat_t, "대분류", "최종판매가", f"{sel_brand} 카테고리 비중 (매출)", cmap=_CAT_CMAP),
                    use_container_width=True)
with cc2:
    ct = cat_t.copy()
    ct.insert(0, "Rank", np.arange(1, len(ct) + 1))
    ct = ct[["Rank", "대분류", "수량", "최종판매가", "객단가", "수익률", "매출비중"]]
    _ftc, _fcc = format_table(ct)
    render_table(_ftc, hide_index=True, use_container_width=True, height=60 + len(ct) * 36,
                 column_config={**_fcc, "Rank": st.column_config.TextColumn("#")})

# 8개 대분류로 매핑 안 된(미분류) 원본 카테고리 표시 — 매핑 추가 참고용
if "대분류_원본" in f.columns and (f["대분류"] == "미분류").any():
    _miss = (f.loc[f["대분류"] == "미분류", "대분류_원본"].astype(str)
             .value_counts().head(12).index.tolist())
    _miss = [m for m in _miss if m not in ("미분류", "nan", "")]
    if _miss:
        st.caption(f"⚠️ 8개 대분류로 매핑 안 된 항목 → '미분류'로 모음: {', '.join(_miss)}  ·  어느 대분류로 넣을지 알려주면 추가합니다.")

# 대분류 × 시즌 구성 (최근 시즌, 누적)
if len(recent_seasons) >= 2:
    gm = g[g["시즌"].isin(recent_seasons)]
    cs = gm.groupby(["시즌", "대분류"], as_index=False)["최종판매가"].sum()
    cs["시즌"] = pd.Categorical(cs["시즌"], categories=recent_seasons, ordered=True)
    cs = cs.sort_values("시즌")
    fig_cs = px.bar(cs, x="시즌", y="최종판매가", color="대분류", barmode="stack",
                    title=f"{sel_brand} 시즌별 카테고리 구성",
                    category_orders={"시즌": recent_seasons, "대분류": _cat_order_all},
                    color_discrete_map=_CAT_CMAP,
                    labels={"최종판매가": "매출"})
    fig_cs.update_layout(xaxis_type="category", margin=dict(t=54, b=10), legend_title_text="대분류")
    st.plotly_chart(fig_cs, use_container_width=True)

# =============================================================
# 4) 브랜드 TOP 10 상품 (3분할: 매출액 / 수익 / 수익률)
# =============================================================
st.markdown(f"<div class='section-title'>{sel_brand} TOP 10 상품</div>", unsafe_allow_html=True)
prod = aggregate(f, ["브랜드", "대분류", "라인명"], metric_cols).reset_index(drop=True)
# 라인별 대표 시즌 = 그 라인 매출이 가장 큰 시즌 (입고일 기준 시즌)
if not prod.empty and "시즌" in f.columns:
    _lseason = (f.groupby(["라인명", "시즌"])["최종판매가"].sum().reset_index()
                .sort_values("최종판매가", ascending=False)
                .drop_duplicates("라인명").set_index("라인명")["시즌"])
    prod["시즌"] = prod["라인명"].map(_lseason).fillna("미상")
if prod.empty:
    st.caption("표시할 상품이 없습니다.")
else:
    _by_sales = prod.sort_values("최종판매가", ascending=False).head(10).reset_index(drop=True)
    _by_profit = prod.sort_values("수익원(실배송비)", ascending=False).head(10).reset_index(drop=True)
    _by_rate = (prod[(prod["최종판매가"] > 0) & (prod["수량"] >= 5)].sort_values("수익률", ascending=False)
                .head(10).reset_index(drop=True))
    _t1, _t2, _t3 = st.columns(3)
    with _t1:
        st.markdown("**매출액순**")
        st.markdown(product_cards_html(_by_sales, n=10, img_px=72), unsafe_allow_html=True)
    with _t2:
        st.markdown("**수익순**")
        st.markdown(product_cards_html(_by_profit, n=10, img_px=72), unsafe_allow_html=True)
    with _t3:
        st.markdown("**수익률순** (5개 이상 판매)")
        st.markdown(product_cards_html(_by_rate, n=10, img_px=72), unsafe_allow_html=True)

# =============================================================
# =============================================================
# 6) 자동 요약
# =============================================================
st.markdown("<div class='section-title'>자동 요약</div>", unsafe_allow_html=True)
_mt = aggregate(f, ["쇼핑몰"], metric_cols).head(1)
_ct = aggregate(f, ["대분류"], metric_cols).head(1)
_lt = aggregate(f, ["라인명"], metric_cols).head(1)
sm = []
sm.append(f"- **{sel_brand}** 선택 시즌 합계 매출 **{eok(tot_sales)}** · 수량 **{num(tot_qty)}개** · "
          f"객단가 **{eok(avg_price)}** · 수익률 **{pct(profit_rate)}**.")
if latest_season:
    if has_prev:
        sm.append(f"- 최근 시즌 **{latest_season}** 매출 **{eok(ls_sales)}**, "
                  f"전년 동시즌({prev_same}) **{eok(ps_sales)}** 대비 **{growth_pct(season_yoy)}**.")
    else:
        sm.append(f"- 최근 시즌 **{latest_season}** 매출 **{eok(ls_sales)}** (전년 동시즌 데이터 없음).")
if not _mt.empty:
    sm.append(f"- 쇼핑몰 1위 **{_mt.iloc[0]['쇼핑몰']}** · {eok(_mt.iloc[0]['최종판매가'])} "
              f"(비중 {pct(_mt.iloc[0].get('매출비중', np.nan))}).")
if not _ct.empty:
    sm.append(f"- 카테고리 1위 **{_ct.iloc[0]['대분류']}** · {eok(_ct.iloc[0]['최종판매가'])}.")
if not _lt.empty:
    sm.append(f"- 베스트 라인 **{_lt.iloc[0]['라인명']}** · {eok(_lt.iloc[0]['최종판매가'])}.")
st.markdown("\n".join(sm))

# =============================================================
# 7) 재고 (이 브랜드 · 재고 파일이 있을 때만)
# =============================================================
if "stock_df" in dir() and isinstance(stock_df, pd.DataFrame) and not stock_df.empty:
    bstock = stock_df[stock_df["브랜드"].astype(str).str.strip() == str(sel_brand).strip()].copy().reset_index(drop=True)
    st.markdown(f"<div class='section-title'>📦 재고 — {sel_brand}</div>", unsafe_allow_html=True)
    if bstock.empty:
        st.caption(f"재고 파일에서 '{sel_brand}' 브랜드를 찾지 못했습니다. (판매 데이터와 재고의 브랜드 표기가 다를 수 있음)")
    else:
        bstock["수량"] = pd.to_numeric(bstock["수량"], errors="coerce").fillna(0)
        bstock["총원가"] = pd.to_numeric(bstock["총원가"], errors="coerce").fillna(0)
        # 대분류: 재고 자체 분류(매출과 동일 _CATEGORY_MAP 적용됨) 우선, '미분류'만 판매 라인매핑으로 보강
        if "대분류" not in bstock.columns:
            bstock["대분류"] = "미분류"
        bstock["대분류"] = (bstock["대분류"].astype(str).str.strip()
                          .replace({"": "미분류", "nan": "미분류", "None": "미분류"}))
        _mask = bstock["대분류"] == "미분류"
        if _mask.any():
            _line2cat = (df.groupby("라인명")["대분류"]
                         .agg(lambda s: s.mode().iat[0] if len(s.mode()) else np.nan))
            _bf = bstock.loc[_mask, "라인명"].astype(str).str.strip().map(_line2cat)
            _bf = _bf.where(_bf.isin(_ALLOWED_CATS))
            bstock.loc[_mask, "대분류"] = _bf.fillna("미분류").values
        # 시즌: 재고 입고일자 기준
        bstock["시즌"] = _mk_season(bstock["입고일자"]) if "입고일자" in bstock.columns else "미상"
        # 재고용 대분류 필터 적용
        bstock = bstock[bstock["대분류"].isin(sel_cats_stock)].reset_index(drop=True)
        if bstock.empty:
            st.caption("선택한 대분류에 해당하는 재고가 없습니다.")

        b1, b2, b3 = st.columns(3)
        b1.metric("총 재고수량", f"{int(bstock['수량'].sum()):,}개")
        b2.metric("총 재고원가", eok(bstock["총원가"].sum()))
        b3.metric("라인 수", f"{bstock['라인명'].nunique():,}")
        # 8개로 매핑 안 된(미분류) 재고 카테고리 표시 — 키워드 추가 참고용
        _src_col = "대분류_원본" if "대분류_원본" in bstock.columns else "카테고리"
        if _src_col in bstock.columns and (bstock["대분류"] == "미분류").any():
            _ms = (bstock.loc[bstock["대분류"] == "미분류", _src_col].astype(str)
                   .value_counts().head(15).index.tolist())
            _ms = [m for m in _ms if m not in ("미분류", "nan", "")]
            if _ms:
                st.caption(f"⚠️ 8개 대분류로 매핑 안 된 재고 카테고리: {', '.join(_ms)}  ·  어느 대분류인지 알려주면 추가합니다.")
        with st.expander("🔧 재고 분류 진단 — 어느 컬럼을 분류로 읽었나"):
            if _STOCK_CAT_DEBUG:
                st.write(f"분류로 채택한 컬럼: **{_STOCK_CAT_DEBUG.get('selected')}** "
                         f"(8개 매핑률 {_STOCK_CAT_DEBUG.get('rate')})")
                st.write("후보 컬럼별 매핑률:", _STOCK_CAT_DEBUG.get("candidates"))
                st.caption(f"재고 전체 컬럼: {', '.join(_STOCK_CAT_DEBUG.get('all_cols', []))}")
            if "대분류_원본" in bstock.columns:
                _src = (bstock.groupby("대분류_원본")["대분류"].first()
                        .reset_index().rename(columns={"대분류_원본": "재고 원본값", "대분류": "→ 매핑"}))
                render_table(_src, hide_index=True, use_container_width=True,
                             height=min(60 + len(_src) * 32, 360))

        # ---- 회전율 · 완판 분석 (라인 × 시즌) ----
        st.markdown("**회전율 · 완판 분석** (라인 × 시즌)")
        _tdy = pd.Timestamp.now().normalize()
        # (라인, 시즌)별 입고량 + 첫입고일 — 재고 입고이벤트에서
        _inb_rows = []
        if "입고이벤트" in bstock.columns:
            for _ln, _evs in zip(bstock["라인명"].astype(str).str.strip(), bstock["입고이벤트"]):
                if not isinstance(_evs, list):
                    continue
                for _n, _qq in _evs:
                    _din = _tdy - pd.Timedelta(days=int(_n))
                    _slab = _mk_one_season(_din)
                    if _slab == "미상" or _season_meta(_slab)[1] < SEASON_MIN_YEAR:
                        continue
                    _inb_rows.append((_ln, _slab, int(_qq), _din))
        if not _inb_rows:
            st.caption("입고이력(I열 'N일전/수량')이 없어 회전율을 계산할 수 없습니다.")
        else:
            _inb_ls = (pd.DataFrame(_inb_rows, columns=["라인명", "시즌", "입고량", "입고일"])
                       .groupby(["라인명", "시즌"]).agg(입고량=("입고량", "sum"),
                                                      첫입고일=("입고일", "min")).reset_index())
            _gk = g["라인명"].astype(str).str.strip()
            _gg = g.assign(_ln=_gk)
            _gg["_정산"] = (pd.to_numeric(_gg["정산금"], errors="coerce").fillna(0)
                          if "정산금" in _gg.columns else 0)
            _sal_ls = (_gg.groupby(["_ln", "시즌"]).agg(
                판매량=("수량", "sum"), 매출=("최종판매가", "sum"),
                수익=("수익원(실배송비)", "sum"), 정산금합=("_정산", "sum")
            ).reset_index().rename(columns={"_ln": "라인명"}))
            _last_sale = _gg.groupby("_ln")["날짜"].max()
            _ln_cat = bstock.groupby(bstock["라인명"].astype(str).str.strip())["대분류"].first()
            _rt = _inb_ls.merge(_sal_ls, on=["라인명", "시즌"], how="left")
            _rt["판매량"] = _rt["판매량"].fillna(0)
            for _cc in ("매출", "수익", "정산금합"):
                if _cc in _rt.columns:
                    _rt[_cc] = _rt[_cc].fillna(0)
            _rt = _rt[_rt["입고량"] > 0].copy()
            _rt["회전율"] = _rt["판매량"] / _rt["입고량"] * 100
            _rt["현재고"] = (_rt["입고량"] - _rt["판매량"]).clip(lower=0)
            _rt["마지막판매"] = _rt["라인명"].map(_last_sale)
            _rt["완판기간"] = (_rt["마지막판매"] - _rt["첫입고일"]).dt.days
            _rt["입고경과일"] = (_tdy - _rt["첫입고일"]).dt.days
            _rt["수익률"] = np.where(_rt.get("매출", 0) > 0, _rt.get("수익", 0) / _rt["매출"].replace(0, np.nan) * 100, np.nan)
            _rt["평균정산금"] = np.where(_rt["판매량"] > 0, _rt.get("정산금합", 0) / _rt["판매량"].replace(0, np.nan), np.nan)
            _rt["대분류"] = _rt["라인명"].map(_ln_cat)

            def _turn_val(r):
                _el = (f' · 입고 {int(r["입고경과일"])}일' if pd.notna(r.get("입고경과일")) else '')
                _st = (f' · 평균정산 {eok(r["평균정산금"])}원' if pd.notna(r.get("평균정산금")) else '')
                return (f'<div style="font-size:13px;color:#0f172a;">회전율 '
                        f'<span style="color:#10b981;font-weight:700;">{r["회전율"]:.1f}%</span> '
                        f'<span style="color:#64748b;">· 입고 {int(r["입고량"]):,} · 현재고 {int(r["현재고"]):,} '
                        f'· 판매 {int(round(r["판매량"])):,}{_el}{_st}</span></div>')

            def _sellout_val(r):
                _fi = pd.Timestamp(r["첫입고일"]).strftime("%Y.%m.%d") if pd.notna(r["첫입고일"]) else "?"
                _ls = pd.Timestamp(r["마지막판매"]).strftime("%Y.%m.%d") if pd.notna(r["마지막판매"]) else "?"
                _pr = f' · 수익률 {r["수익률"]:.1f}%' if pd.notna(r.get("수익률")) else ''
                _st = f' · 평균정산 {eok(r["평균정산금"])}원' if pd.notna(r.get("평균정산금")) else ''
                return (f'<div style="font-size:13px;color:#0f172a;">'
                        f'<span style="color:#b45309;font-weight:700;">완판 {int(r["완판기간"])}일</span> '
                        f'<span style="color:#64748b;">· 입고 {int(r["입고량"]):,} · 판매 {int(round(r["판매량"])):,}{_pr}{_st}</span></div>'
                        f'<div style="font-size:12px;color:#64748b;">최초입고 {_fi} → 최종판매 {_ls}</div>')

            # 3분할: 회전율 낮은 TOP50 / 회전율 높은 TOP50 / 완판 TOP50
            # (완판 판정: 부동소수 오차로 100.0이 미완판으로 새는 것 방지 → 99.95% 이상은 완판)
            _SOLD = _rt["회전율"] >= 99.95
            _live = _rt[~_SOLD]
            # 회전율 낮은: 입고 30일 이내(아직 팔릴 시간 부족)는 제외
            _tp_low = (_live[_live["입고경과일"] > 30].sort_values("회전율")
                       .head(50).reset_index(drop=True))
            _tp_high = _live.sort_values("회전율", ascending=False).head(50).reset_index(drop=True)
            _sd = (_rt[_SOLD & _rt["완판기간"].notna() & (_rt["완판기간"] >= 0)]
                   .sort_values("완판기간").head(50).reset_index(drop=True))

            _c1, _c2, _c3 = st.columns(3)
            with _c1:
                st.markdown("**회전율 낮은 TOP50** (입고 30일↑ · 안 팔림)")
                if _tp_low.empty:
                    st.caption("대상 없음")
                else:
                    st.markdown(metric_cards_html(_tp_low, _turn_val, n=len(_tp_low), img_px=92, start=1, step=1),
                                unsafe_allow_html=True)
            with _c2:
                st.markdown("**회전율 높은 TOP50** (완판 100% 제외)")
                if _tp_high.empty:
                    st.caption("대상 없음")
                else:
                    st.markdown(metric_cards_html(_tp_high, _turn_val, n=len(_tp_high), img_px=92, start=1, step=1),
                                unsafe_allow_html=True)
            with _c3:
                st.markdown("**완판 TOP50** (100% · 빨리 완판순)")
                if _sd.empty:
                    st.caption("완판(100% 회전) 상품 없음")
                else:
                    st.markdown(metric_cards_html(_sd, _sellout_val, n=len(_sd), img_px=92, start=1, step=1),
                                unsafe_allow_html=True)
            st.caption("회전율 = 그 시즌 판매량 ÷ 그 시즌 입고량 (라인×시즌). 같은 라인도 입고 시즌이 다르면 별도. "
                       "완판=그 시즌 100% 소진(99.95%↑), 완판기간=시즌 첫 입고일~라인 최종 판매일.")

        # 총재고 순위 (회전율 밑 · 판매 카드 스타일 · 총원가순)
        st.markdown("**총재고 순위**")
        _has_elapsed = "입고경과일행" in bstock.columns and bstock["입고경과일행"].notna().any()
        _agg_kw = dict(총원가=("총원가", "sum"), 재고수량=("수량", "sum"),
                       브랜드=("브랜드", "first"), 대분류=("대분류", "first"), 시즌=("시즌", "first"))
        if _has_elapsed:
            _agg_kw["입고경과일"] = ("입고경과일행", "max")  # 모든 사이즈 중 가장 오래된(=일수 최대)
        bt = (bstock.groupby("라인명", dropna=False).agg(**_agg_kw)
              .reset_index().sort_values("총원가", ascending=False).head(10).reset_index(drop=True))
        if bt.empty:
            st.caption("표시할 재고가 없습니다.")
        else:
            _bl = bt.iloc[0::2]
            _br = bt.iloc[1::2]
            kc1, kc2 = st.columns(2)
            with kc1:
                st.markdown(stock_cards_html(_bl, n=len(_bl), img_px=130, start=1, step=2),
                            unsafe_allow_html=True)
            with kc2:
                st.markdown(stock_cards_html(_br, n=len(_br), img_px=130, start=2, step=2),
                            unsafe_allow_html=True)

        sc1, sc2 = st.columns(2)
        # 카테고리(대분류)별 재고 비중 — 원가 기준
        with sc1:
            st.markdown("**카테고리별 재고 비중 (원가)**")
            catg = (bstock.groupby("대분류", dropna=False)["총원가"].sum()
                    .reset_index().sort_values("총원가", ascending=False))
            catg = catg[catg["총원가"] > 0]
            if catg.empty:
                st.caption("재고원가 데이터가 없습니다.")
            else:
                st.plotly_chart(share_donut(catg, "대분류", "총원가", "카테고리별 재고원가"),
                                use_container_width=True)
                _ct = catg["총원가"].sum()
                cdisp = pd.DataFrame({"대분류": catg["대분류"].values,
                                      "재고원가": catg["총원가"].round(0).astype("int64").values,
                                      "비중": (catg["총원가"] / _ct * 100).round(1).values})
                render_table(cdisp, hide_index=True, use_container_width=True,
                             height=min(60 + len(cdisp) * 36, 320),
                             column_config={"재고원가": st.column_config.NumberColumn("재고원가", format="localized"),
                                            "비중": st.column_config.NumberColumn("비중", format="%.1f%%")})
        # 시즌별 재고 비중 — 원가 기준
        with sc2:
            st.markdown("**시즌별 재고 비중 (원가)**")
            seg = bstock.groupby("시즌", dropna=False)["총원가"].sum().reset_index()
            seg = seg[seg["총원가"] > 0]
            if seg.empty:
                st.caption("재고원가 데이터가 없습니다.")
            else:
                def _sk(lbl):
                    try:
                        return int(lbl[2:]) * 10 + (1 if lbl[:2] == "SS" else 2)
                    except Exception:
                        return 9999  # '미상' 맨 뒤
                seg["정렬"] = seg["시즌"].map(_sk)
                seg = seg.sort_values("정렬")
                seg["타입"] = np.where(seg["시즌"].str.startswith("SS"), "SS",
                                     np.where(seg["시즌"].str.startswith("FW"), "FW", "미상"))
                seg["라벨"] = seg["총원가"].apply(eok)
                fig_sg = px.bar(seg, x="시즌", y="총원가", color="타입", text="라벨",
                                category_orders={"시즌": seg["시즌"].tolist()},
                                color_discrete_map={"SS": "#10b981", "FW": "#6366f1", "미상": "#94a3b8"},
                                title="시즌별 재고원가", labels={"총원가": "재고원가"})
                fig_sg.update_traces(textposition="outside", textangle=0, cliponaxis=False)
                fig_sg.update_layout(xaxis_type="category", margin=dict(t=54, b=10),
                                     legend_title_text="", uniformtext_minsize=8, uniformtext_mode="hide")
                fig_sg.update_xaxes(title_text="")
                fig_sg.update_yaxes(title_text="")
                st.plotly_chart(fig_sg, use_container_width=True)
                _st = seg["총원가"].sum()
                sdisp = pd.DataFrame({"시즌": seg["시즌"].values,
                                      "재고원가": seg["총원가"].round(0).astype("int64").values,
                                      "비중": (seg["총원가"] / _st * 100).round(1).values})
                render_table(sdisp, hide_index=True, use_container_width=True,
                             height=min(60 + len(sdisp) * 36, 320),
                             column_config={"재고원가": st.column_config.NumberColumn("재고원가", format="localized"),
                                            "비중": st.column_config.NumberColumn("비중", format="%.1f%%")})

# =============================================================
# 7-2) 이미지 없는 상품 → 터미널(콘솔)에 출력 (화면엔 표로 안 띄움)
# =============================================================
if img_map:
    def _no_img(L):
        return not str(img_map.get(str(L).strip(), "")).strip()

    _sl = aggregate(f, ["브랜드", "대분류", "라인명"], metric_cols)
    _sl_no = _sl[_sl["라인명"].map(_no_img)].reset_index(drop=True)
    _stock_no = None
    if "bstock" in dir() and isinstance(bstock, pd.DataFrame) and not bstock.empty and "총원가" in bstock.columns:
        _kl = bstock.groupby("라인명", dropna=False)["총원가"].sum().reset_index()
        _stock_no = _kl[_kl["라인명"].map(_no_img)].sort_values("총원가", ascending=False).reset_index(drop=True)

    print("\n" + "=" * 72)
    print(f"[이미지 미등록 상품]  브랜드: {sel_brand}")
    print(f"  판매 상품 {len(_sl_no)}개 라인 (이미지 없음, 매출순)")
    for _i, (_ln, _v) in enumerate(zip(_sl_no["라인명"], _sl_no["최종판매가"]), 1):
        try:
            print(f"    {_i:>3}. {_ln}  (매출 {int(_v):,})")
        except Exception:
            print(f"    {_i:>3}. {_ln}")
    if _stock_no is not None:
        print(f"  재고 상품 {len(_stock_no)}개 라인 (이미지 없음, 총원가순)")
        for _i, (_ln, _v) in enumerate(zip(_stock_no["라인명"], _stock_no["총원가"]), 1):
            try:
                print(f"    {_i:>3}. {_ln}  (원가 {int(_v):,})")
            except Exception:
                print(f"    {_i:>3}. {_ln}")
    print("=" * 72 + "\n")

    _msg = f"🚫 이미지 없는 상품 — 판매 {len(_sl_no)}개 라인"
    if _stock_no is not None:
        _msg += f" · 재고 {len(_stock_no)}개 라인"
    st.caption(_msg + " (상세 목록은 앱을 실행한 터미널에 출력됩니다)")
